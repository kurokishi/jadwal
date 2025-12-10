import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import time, timedelta, datetime, date
import io
import re
import json
import os
import shutil
from pathlib import Path
import time as systime
import threading
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ==================== KONFIGURASI SISTEM ====================

# Mode: AUTO (monitoring folder) atau MANUAL (upload manual)
SYSTEM_MODE = "AUTO"  # Ganti ke "MANUAL" untuk mode manual

# Folder untuk auto-monitoring (untuk mode AUTO)
INPUT_FOLDER = "./input_files"
OUTPUT_FOLDER = "./output_files"
ARCHIVE_FOLDER = "./archive"

# Warna untuk styling
FILL_R = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
FILL_E = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
FILL_OVERLIMIT = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Warna untuk Kanban
KANBAN_COLORS = {
    "todo": "#FF6B6B",
    "in_progress": "#FFD93D",
    "review": "#6BCF7F",
    "done": "#4D96FF"
}

# Konfigurasi default
TIME_SLOTS = [
    time(7, 30), time(8, 0), time(8, 30), time(9, 0), time(9, 30),
    time(10, 0), time(10, 30), time(11, 0), time(11, 30), time(12, 0),
    time(12, 30), time(13, 0), time(13, 30), time(14, 0), time(14, 30)
]
TIME_SLOTS_STR = [t.strftime("%H:%M") for t in TIME_SLOTS]

HARI_ORDER = {"Senin": 1, "Selasa": 2, "Rabu": 3, "Kamis": 4, "Jum'at": 5}
HARI_INDONESIA = ["Senin", "Selasa", "Rabu", "Kamis", "Jum'at"]

# ==================== SESSION STATE & DATABASE ====================

class KanbanDatabase:
    """Database sederhana untuk menyimpan data Kanban"""
    
    def __init__(self):
        self.data_file = "kanban_database.json"
        self.load_data()
    
    def load_data(self):
        """Load data dari file JSON"""
        try:
            if os.path.exists(self.data_file):
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    st.session_state.kanban_tasks = data.get('kanban_tasks', {
                        "todo": [], "in_progress": [], "review": [], "done": []
                    })
                    st.session_state.kanban_next_id = data.get('kanban_next_id', 1)
                    st.session_state.system_logs = data.get('system_logs', [])
                    st.session_state.processed_files = data.get('processed_files', [])
            else:
                self.init_default_data()
        except:
            self.init_default_data()
    
    def init_default_data(self):
        """Inisialisasi data default"""
        if 'kanban_tasks' not in st.session_state:
            st.session_state.kanban_tasks = {
                "todo": [], "in_progress": [], "review": [], "done": []
            }
        if 'kanban_next_id' not in st.session_state:
            st.session_state.kanban_next_id = 1
        if 'system_logs' not in st.session_state:
            st.session_state.system_logs = []
        if 'processed_files' not in st.session_state:
            st.session_state.processed_files = []
        if 'auto_mode' not in st.session_state:
            st.session_state.auto_mode = False
        if 'observer_running' not in st.session_state:
            st.session_state.observer_running = False
    
    def save_data(self):
        """Simpan data ke file JSON"""
        data = {
            'kanban_tasks': st.session_state.kanban_tasks,
            'kanban_next_id': st.session_state.kanban_next_id,
            'system_logs': st.session_state.system_logs[-100:],  # Simpan 100 log terakhir
            'processed_files': st.session_state.processed_files[-50:]  # Simpan 50 file terakhir
        }
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    
    def add_log(self, message, log_type="INFO"):
        """Tambahkan log ke sistem"""
        log_entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "type": log_type,
            "message": message
        }
        st.session_state.system_logs.append(log_entry)
        self.save_data()
    
    def add_processed_file(self, filename, status="success"):
        """Catat file yang sudah diproses"""
        file_entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "filename": filename,
            "status": status
        }
        st.session_state.processed_files.append(file_entry)
        self.save_data()

# Inisialisasi database
db = KanbanDatabase()

# ==================== FILE MONITORING SYSTEM ====================

class FileWatcher(FileSystemEventHandler):
    """Monitor folder untuk file baru"""
    
    def on_created(self, event):
        if not event.is_directory and event.src_path.endswith('.xlsx'):
            filename = os.path.basename(event.src_path)
            db.add_log(f"File terdeteksi: {filename}", "SYSTEM")
            
            # Delay sedikit untuk memastikan file selesai ditulis
            systime.sleep(2)
            
            # Proses file
            try:
                process_auto_file(event.src_path)
                db.add_log(f"File berhasil diproses: {filename}", "SUCCESS")
            except Exception as e:
                db.add_log(f"Gagal memproses {filename}: {str(e)}", "ERROR")

def start_file_watcher():
    """Mulai monitoring folder"""
    if not os.path.exists(INPUT_FOLDER):
        os.makedirs(INPUT_FOLDER)
    
    event_handler = FileWatcher()
    observer = Observer()
    observer.schedule(event_handler, INPUT_FOLDER, recursive=False)
    observer.start()
    
    st.session_state.observer_running = True
    db.add_log("File watcher dimulai", "SYSTEM")
    
    return observer

def stop_file_watcher(observer):
    """Hentikan monitoring folder"""
    if observer:
        observer.stop()
        observer.join()
    st.session_state.observer_running = False
    db.add_log("File watcher dihentikan", "SYSTEM")

# ==================== AUTO PROCESSING SYSTEM ====================

def process_auto_file(filepath):
    """Proses file secara otomatis"""
    filename = os.path.basename(filepath)
    
    # Buat task di Kanban
    create_auto_kanban_task(filename, "processing")
    
    # Baca dan proses file
    with open(filepath, 'rb') as f:
        result_buffer = process_file_content(f, filename)
    
    # Simpan hasil
    output_path = os.path.join(OUTPUT_FOLDER, f"hasil_{filename}")
    with open(output_path, 'wb') as f:
        f.write(result_buffer.getvalue())
    
    # Pindah ke archive
    archive_path = os.path.join(ARCHIVE_FOLDER, filename)
    shutil.move(filepath, archive_path)
    
    # Update task status
    update_kanban_task_status(filename, "completed")
    
    db.add_processed_file(filename, "success")
    
    return output_path

def create_auto_kanban_task(filename, status="processing"):
    """Buat task Kanban otomatis"""
    task_id = st.session_state.kanban_next_id
    
    task_data = {
        "id": task_id,
        "title": f"AUTO: Proses {filename}",
        "description": f"File diproses secara otomatis oleh sistem",
        "priority": "high" if "urgent" in filename.lower() else "medium",
        "due_date": datetime.now().strftime("%Y-%m-%d"),
        "assignee": "Sistem Otomatis",
        "tags": ["auto-process", "system", "jadwal"],
        "created_by": "Auto-System",
        "created_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "status": status,
        "filename": filename,
        "type": "auto"
    }
    
    st.session_state.kanban_tasks["in_progress"].append(task_data)
    st.session_state.kanban_next_id += 1
    db.save_data()
    
    return task_id

def update_kanban_task_status(filename, status="completed"):
    """Update status task di Kanban"""
    for column in ["todo", "in_progress", "review", "done"]:
        for task in st.session_state.kanban_tasks[column]:
            if task.get("filename") == filename:
                task["status"] = status
                task["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                
                # Pindah ke Done jika completed
                if status == "completed":
                    st.session_state.kanban_tasks[column].remove(task)
                    st.session_state.kanban_tasks["done"].append(task)
                
                db.save_data()
                return True
    return False

# ==================== CORE PROCESSING FUNCTIONS ====================

def parse_time_range(time_str):
    """Parse rentang waktu"""
    if pd.isna(time_str) or str(time_str).strip() == "":
        return None, None
    
    clean_str = str(time_str).strip().replace(' ', '').replace('.', ':')
    pattern = r'(\d{1,2}:\d{2})-(\d{1,2}:\d{2})'
    match = re.search(pattern, clean_str)
    if not match:
        return None, None
    
    try:
        start_str, end_str = match.groups()
        start_hour, start_minute = map(int, start_str.split(':'))
        end_hour, end_minute = map(int, end_str.split(':'))
        
        start_time = time(start_hour, start_minute)
        end_time = time(end_hour, end_minute)
        
        if end_time < start_time:
            end_time = time(end_hour + 24, end_minute)
        
        return start_time, end_time
    except:
        return None, None

def time_overlap(slot_start, slot_end, schedule_start, schedule_end):
    """Cek overlap waktu"""
    if schedule_start is None or schedule_end is None:
        return False
    return not (slot_end <= schedule_start or slot_start >= schedule_end)

def process_schedule(df, jenis_poli):
    """Proses dataframe jadwal"""
    results = []
    
    for (dokter, poli_asal), group in df.groupby(['Nama Dokter', 'Poli Asal']):
        hari_schedules = {}
        
        for hari in HARI_INDONESIA:
            if hari not in group.columns:
                continue
                
            time_ranges = []
            for time_str in group[hari]:
                start_time, end_time = parse_time_range(time_str)
                if start_time and end_time:
                    if end_time > time(14, 30):
                        end_time = time(14, 30)
                    time_ranges.append((start_time, end_time))
            
            hari_schedules[hari] = time_ranges
        
        for hari in HARI_INDONESIA:
            if hari not in hari_schedules or not hari_schedules[hari]:
                continue
                
            merged_ranges = []
            for start, end in sorted(hari_schedules[hari]):
                if not merged_ranges:
                    merged_ranges.append([start, end])
                else:
                    last_start, last_end = merged_ranges[-1]
                    if start <= last_end:
                        merged_ranges[-1][1] = max(last_end, end)
                    else:
                        merged_ranges.append([start, end])
            
            row = {
                'POLI ASAL': poli_asal,
                'JENIS POLI': jenis_poli,
                'HARI': hari,
                'DOKTER': dokter
            }
            
            for i, slot_time in enumerate(TIME_SLOTS):
                slot_start = slot_time
                slot_end = (datetime.combine(datetime.today(), slot_start) + 
                           timedelta(minutes=30)).time()
                
                has_overlap = any(
                    time_overlap(slot_start, slot_end, start, end)
                    for start, end in merged_ranges
                )
                
                if has_overlap:
                    row[TIME_SLOTS_STR[i]] = 'R' if jenis_poli == 'Reguler' else 'E'
                else:
                    row[TIME_SLOTS_STR[i]] = ''
            
            results.append(row)
    
    return pd.DataFrame(results)

def apply_styles(ws, max_row):
    """Terapkan styling ke sheet"""
    e_counts = {hari: {slot: 0 for slot in TIME_SLOTS_STR} for hari in HARI_INDONESIA}
    
    for row in range(2, max_row + 1):
        hari = ws.cell(row=row, column=3).value
        if hari not in HARI_INDONESIA:
            continue
            
        for col_idx, slot in enumerate(TIME_SLOTS_STR, start=5):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value == 'E':
                e_counts[hari][slot] += 1
    
    for row in range(2, max_row + 1):
        hari = ws.cell(row=row, column=3).value
        if hari not in HARI_INDONESIA:
            continue
            
        for col_idx, slot in enumerate(TIME_SLOTS_STR, start=5):
            cell = ws.cell(row=row, column=col_idx)
            
            if cell.value == 'R':
                cell.fill = FILL_R
            elif cell.value == 'E':
                cell.fill = FILL_E
                
                if e_counts[hari][slot] > 7:
                    e_rows_for_slot = []
                    for r in range(2, max_row + 1):
                        if (ws.cell(row=r, column=3).value == hari and 
                            ws.cell(row=r, column=col_idx).value == 'E'):
                            e_rows_for_slot.append(r)
                    
                    if len(e_rows_for_slot) > 7:
                        if row in e_rows_for_slot[7:]:
                            cell.fill = FILL_OVERLIMIT

def process_file_content(file_content, filename):
    """Proses konten file"""
    # Baca file
    wb = load_workbook(file_content)
    
    # Cek sheet
    required_sheets = ['Reguler', 'Poleks', 'Jadwal']
    missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
    
    if missing_sheets:
        raise ValueError(f"Sheet hilang: {', '.join(missing_sheets)}")
    
    # Baca data
    file_content.seek(0)
    df_reguler = pd.read_excel(file_content, sheet_name='Reguler')
    file_content.seek(0)
    df_poleks = pd.read_excel(file_content, sheet_name='Poleks')
    
    # Proses jadwal
    df_reguler_processed = process_schedule(df_reguler, 'Reguler')
    df_poleks_processed = process_schedule(df_poleks, 'Poleks')
    
    # Gabungkan dan urutkan
    df_jadwal = pd.concat([df_reguler_processed, df_poleks_processed], ignore_index=True)
    df_jadwal['HARI_ORDER'] = df_jadwal['HARI'].map(HARI_ORDER)
    df_jadwal = df_jadwal.sort_values(['POLI ASAL', 'HARI_ORDER', 'DOKTER', 'JENIS POLI'])
    df_jadwal = df_jadwal.drop('HARI_ORDER', axis=1)
    df_jadwal = df_jadwal.reset_index(drop=True)
    
    # Simpan ke session
    st.session_state.processed_data = df_jadwal
    
    # Buat workbook baru
    new_wb = load_workbook(file_content)
    
    # Update sheet Jadwal
    if 'Jadwal' in new_wb.sheetnames:
        new_wb.remove(new_wb['Jadwal'])
    
    ws_jadwal = new_wb.create_sheet('Jadwal')
    headers = ['POLI ASAL', 'JENIS POLI', 'HARI', 'DOKTER'] + TIME_SLOTS_STR
    
    for col_idx, header in enumerate(headers, start=1):
        ws_jadwal.cell(row=1, column=col_idx, value=header)
    
    for row_idx, row_data in enumerate(df_jadwal.to_dict('records'), start=2):
        ws_jadwal.cell(row=row_idx, column=1, value=row_data['POLI ASAL'])
        ws_jadwal.cell(row=row_idx, column=2, value=row_data['JENIS POLI'])
        ws_jadwal.cell(row=row_idx, column=3, value=row_data['HARI'])
        ws_jadwal.cell(row=row_idx, column=4, value=row_data['DOKTER'])
        
        for col_idx, slot in enumerate(TIME_SLOTS_STR, start=5):
            ws_jadwal.cell(row=row_idx, column=col_idx, value=row_data.get(slot, ''))
    
    apply_styles(ws_jadwal, len(df_jadwal) + 1)
    
    # Simpan ke buffer
    result_buffer = io.BytesIO()
    new_wb.save(result_buffer)
    result_buffer.seek(0)
    
    return result_buffer

# ==================== UI COMPONENTS ====================

def render_dashboard():
    """Dashboard utama sistem otomatis"""
    st.title("🏥 SISTEM OTOMATIS JADWAL POLI")
    st.markdown("### 📊 Dashboard Monitoring Real-time")
    
    # Status sistem
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        auto_status = "🟢 AKTIF" if st.session_state.auto_mode else "🔴 NON-AKTIF"
        st.metric("Mode Otomatis", auto_status)
    
    with col2:
        total_tasks = sum(len(tasks) for tasks in st.session_state.kanban_tasks.values())
        st.metric("Total Tasks", total_tasks)
    
    with col3:
        pending = len(st.session_state.kanban_tasks["todo"]) + len(st.session_state.kanban_tasks["in_progress"])
        st.metric("Pending Tasks", pending)
    
    with col4:
        processed_count = len(st.session_state.processed_files)
        st.metric("Files Diproses", processed_count)
    
    # Kontrol sistem
    st.markdown("---")
    col_control1, col_control2, col_control3 = st.columns(3)
    
    with col_control1:
        if st.button("🚀 START AUTO MODE", use_container_width=True):
            st.session_state.auto_mode = True
            db.add_log("Mode otomatis diaktifkan", "SYSTEM")
            st.rerun()
    
    with col_control2:
        if st.button("⏸️ PAUSE AUTO MODE", use_container_width=True):
            st.session_state.auto_mode = False
            db.add_log("Mode otomatis dijeda", "SYSTEM")
            st.rerun()
    
    with col_control3:
        if st.button("🔄 RESET SYSTEM", use_container_width=True, type="secondary"):
            st.session_state.auto_mode = False
            st.session_state.kanban_tasks = {"todo": [], "in_progress": [], "review": [], "done": []}
            st.session_state.kanban_next_id = 1
            st.session_state.system_logs = []
            st.session_state.processed_files = []
            db.save_data()
            st.rerun()
    
    # Quick actions
    st.markdown("### ⚡ Quick Actions")
    
    col_q1, col_q2, col_q3 = st.columns(3)
    
    with col_q1:
        if st.button("📁 Buka Input Folder", use_container_width=True):
            if os.path.exists(INPUT_FOLDER):
                st.info(f"Folder: {os.path.abspath(INPUT_FOLDER)}")
                files = os.listdir(INPUT_FOLDER)
                if files:
                    st.write("**Files dalam folder:**")
                    for f in files:
                        st.write(f"- {f}")
                else:
                    st.write("Folder kosong")
    
    with col_q2:
        if st.button("📤 Process All Pending", use_container_width=True):
            process_pending_files()
    
    with col_q3:
        if st.button("📥 Download All Results", use_container_width=True):
            download_all_results()
    
    # Tabs untuk berbagai view
    tab1, tab2, tab3, tab4 = st.tabs(["📋 Kanban Board", "📁 File Monitor", "📊 System Logs", "⚙️ Settings"])
    
    with tab1:
        render_kanban_board()
    
    with tab2:
        render_file_monitor()
    
    with tab3:
        render_system_logs()
    
    with tab4:
        render_system_settings()

def render_kanban_board():
    """Render Kanban board yang diperbarui"""
    st.subheader("📋 Kanban Board - Auto System")
    
    # Statistics
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    with col_stat1:
        st.metric("Todo", len(st.session_state.kanban_tasks["todo"]))
    with col_stat2:
        st.metric("In Progress", len(st.session_state.kanban_tasks["in_progress"]))
    with col_stat3:
        st.metric("Review", len(st.session_state.kanban_tasks["review"]))
    with col_stat4:
        st.metric("Done", len(st.session_state.kanban_tasks["done"]))
    
    # Filter
    st.markdown("---")
    col_filter1, col_filter2 = st.columns(2)
    with col_filter1:
        filter_type = st.selectbox("Filter Type", ["All", "Auto", "Manual"], key="filter_type")
    with col_filter2:
        filter_status = st.selectbox("Filter Status", ["All", "Processing", "Completed", "Error"], key="filter_status")
    
    # Columns
    columns = st.columns(4)
    column_data = [
        ("Todo", "todo", KANBAN_COLORS["todo"]),
        ("In Progress", "in_progress", KANBAN_COLORS["in_progress"]),
        ("Review", "review", KANBAN_COLORS["review"]),
        ("Done", "done", KANBAN_COLORS["done"])
    ]
    
    for col, (col_name, col_key, col_color) in zip(columns, column_data):
        with col:
            # Header
            st.markdown(
                f"""
                <div style='
                    background-color: {col_color};
                    padding: 10px;
                    border-radius: 5px;
                    margin-bottom: 10px;
                    color: white;
                    font-weight: bold;
                    text-align: center;
                '>
                {col_name} ({len(st.session_state.kanban_tasks[col_key])})
                </div>
                """,
                unsafe_allow_html=True
            )
            
            # Tasks
            tasks = st.session_state.kanban_tasks[col_key]
            
            # Apply filters
            filtered_tasks = tasks
            if filter_type != "All":
                filtered_tasks = [t for t in filtered_tasks if t.get("type", "manual") == filter_type.lower()]
            if filter_status != "All":
                filtered_tasks = [t for t in filtered_tasks if t.get("status", "") == filter_status.lower()]
            
            if not filtered_tasks:
                st.info("No tasks")
            else:
                for task in filtered_tasks:
                    render_auto_task_card(task, col_key)

def render_auto_task_card(task, current_column):
    """Render task card untuk sistem otomatis"""
    priority_colors = {"high": "#FF0000", "medium": "#FFA500", "low": "#008000"}
    priority = task.get("priority", "medium")
    
    with st.container():
        # Card header
        card_color = priority_colors.get(priority, "#FFA500")
        
        st.markdown(
            f"""
            <div style='
                background-color: white;
                border: 3px solid {card_color};
                border-radius: 8px;
                padding: 12px;
                margin-bottom: 10px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            '>
            <div style='display: flex; justify-content: space-between;'>
                <div>
                    <strong>#{task.get('id')}</strong>
                    <span style='background-color: #e0e0e0; padding: 2px 6px; border-radius: 10px; 
                          font-size: 0.8em; margin-left: 5px;'>
                        {task.get('type', 'manual').upper()}
                    </span>
                </div>
                <small style='color: {card_color}; font-weight: bold;'>{priority.upper()}</small>
            </div>
            <h4 style='margin: 8px 0;'>{task.get('title', 'No Title')}</h4>
            """,
            unsafe_allow_html=True
        )
        
        # Info
        if task.get("description"):
            st.caption(task.get("description"))
        
        if task.get("filename"):
            st.caption(f"📄 {task.get('filename')}")
        
        # Status badge
        status = task.get("status", "")
        status_color = "#4CAF50" if status == "completed" else "#FF9800" if status == "processing" else "#F44336"
        if status:
            st.markdown(
                f"<div style='background-color: {status_color}; color: white; padding: 4px 8px; "
                f"border-radius: 5px; display: inline-block; margin: 5px 0;'>"
                f"{status.upper()}</div>",
                unsafe_allow_html=True
            )
        
        # Details
        col_info1, col_info2 = st.columns(2)
        with col_info1:
            if task.get("assignee"):
                st.caption(f"👤 {task.get('assignee')}")
        with col_info2:
            if task.get("created_date"):
                st.caption(f"🕐 {task.get('created_date')}")
        
        st.markdown("</div>", unsafe_allow_html=True)
        
        # Actions
        if current_column != "done":
            col_act1, col_act2 = st.columns(2)
            with col_act1:
                if st.button("✅ Complete", key=f"complete_{task['id']}", use_container_width=True):
                    complete_task(task['id'], current_column)
            with col_act2:
                if st.button("🗑️ Delete", key=f"delete_{task['id']}", use_container_width=True):
                    delete_auto_task(task['id'], current_column)

def complete_task(task_id, current_column):
    """Tandai task sebagai selesai"""
    for task in st.session_state.kanban_tasks[current_column]:
        if task["id"] == task_id:
            task["status"] = "completed"
            task["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            st.session_state.kanban_tasks[current_column].remove(task)
            st.session_state.kanban_tasks["done"].append(task)
            db.save_data()
            db.add_log(f"Task #{task_id} completed", "SUCCESS")
            break
    st.rerun()

def delete_auto_task(task_id, current_column):
    """Hapus task"""
    st.session_state.kanban_tasks[current_column] = [
        t for t in st.session_state.kanban_tasks[current_column] if t["id"] != task_id
    ]
    db.save_data()
    db.add_log(f"Task #{task_id} deleted", "INFO")
    st.rerun()

def render_file_monitor():
    """Monitor file system"""
    st.subheader("📁 File System Monitor")
    
    # Create folders if not exist
    for folder in [INPUT_FOLDER, OUTPUT_FOLDER, ARCHIVE_FOLDER]:
        if not os.path.exists(folder):
            os.makedirs(folder)
    
    # File upload untuk mode manual
    st.markdown("### 📤 Upload File Manual")
    uploaded_file = st.file_uploader("Upload file Excel", type=['xlsx'], key="manual_upload")
    
    if uploaded_file:
        col_up1, col_up2 = st.columns(2)
        with col_up1:
            if st.button("🚀 Process File", use_container_width=True):
                with st.spinner("Memproses..."):
                    try:
                        result = process_file_content(uploaded_file, uploaded_file.name)
                        
                        # Save to output
                        output_path = os.path.join(OUTPUT_FOLDER, f"hasil_{uploaded_file.name}")
                        with open(output_path, 'wb') as f:
                            f.write(result.getvalue())
                        
                        # Create Kanban task
                        create_manual_kanban_task(uploaded_file.name, "completed")
                        
                        st.success(f"✅ File diproses! Hasil disimpan di: {output_path}")
                        db.add_log(f"Manual upload processed: {uploaded_file.name}", "SUCCESS")
                        
                        # Download button
                        st.download_button(
                            label="📥 Download Hasil",
                            data=result,
                            file_name=f"hasil_{uploaded_file.name}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                        db.add_log(f"Manual upload failed: {str(e)}", "ERROR")
        
        with col_up2:
            if st.button("➕ Add to Queue", use_container_width=True):
                create_manual_kanban_task(uploaded_file.name, "pending")
                st.success("Task ditambahkan ke queue!")
    
    # File system status
    st.markdown("---")
    st.subheader("📊 File System Status")
    
    col_fs1, col_fs2, col_fs3 = st.columns(3)
    
    with col_fs1:
        st.metric("Input Files", len(os.listdir(INPUT_FOLDER)) if os.path.exists(INPUT_FOLDER) else 0)
        if st.button("View Input Folder"):
            show_folder_contents(INPUT_FOLDER)
    
    with col_fs2:
        st.metric("Output Files", len(os.listdir(OUTPUT_FOLDER)) if os.path.exists(OUTPUT_FOLDER) else 0)
        if st.button("View Output Folder"):
            show_folder_contents(OUTPUT_FOLDER)
    
    with col_fs3:
        st.metric("Archive Files", len(os.listdir(ARCHIVE_FOLDER)) if os.path.exists(ARCHIVE_FOLDER) else 0)
        if st.button("View Archive"):
            show_folder_contents(ARCHIVE_FOLDER)
    
    # Recent processed files
    st.markdown("---")
    st.subheader("📋 Recently Processed Files")
    
    if st.session_state.processed_files:
        recent_files = st.session_state.processed_files[-10:]  # Last 10 files
        for file_info in reversed(recent_files):
            status_icon = "✅" if file_info["status"] == "success" else "❌"
            st.write(f"{status_icon} **{file_info['filename']}** - {file_info['timestamp']}")
    else:
        st.info("No files processed yet")

def create_manual_kanban_task(filename, status="pending"):
    """Buat task manual di Kanban"""
    task_id = st.session_state.kanban_next_id
    
    task_data = {
        "id": task_id,
        "title": f"MANUAL: {filename}",
        "description": f"File diupload secara manual",
        "priority": "medium",
        "due_date": datetime.now().strftime("%Y-%m-%d"),
        "assignee": "User",
        "tags": ["manual-upload", "user"],
        "created_by": "User",
        "created_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "status": status,
        "filename": filename,
        "type": "manual"
    }
    
    st.session_state.kanban_tasks["todo"].append(task_data)
    st.session_state.kanban_next_id += 1
    db.save_data()
    
    return task_id

def show_folder_contents(folder_path):
    """Tampilkan konten folder"""
    if os.path.exists(folder_path):
        files = os.listdir(folder_path)
        if files:
            st.write(f"**Files in {folder_path}:**")
            for file in files:
                file_path = os.path.join(folder_path, file)
                file_size = os.path.getsize(file_path) / 1024  # KB
                modified = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime("%Y-%m-%d %H:%M")
                st.write(f"- {file} ({file_size:.1f} KB, modified: {modified})")
        else:
            st.info("Folder kosong")
    else:
        st.error("Folder tidak ditemukan")

def process_pending_files():
    """Proses semua file yang pending"""
    if os.path.exists(INPUT_FOLDER):
        files = [f for f in os.listdir(INPUT_FOLDER) if f.endswith('.xlsx')]
        
        if not files:
            st.info("Tidak ada file pending")
            return
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, filename in enumerate(files):
            status_text.text(f"Memproses {filename} ({i+1}/{len(files)})")
            filepath = os.path.join(INPUT_FOLDER, filename)
            
            try:
                process_auto_file(filepath)
                st.success(f"✅ {filename} diproses")
            except Exception as e:
                st.error(f"❌ Gagal memproses {filename}: {str(e)}")
            
            progress_bar.progress((i + 1) / len(files))
        
        status_text.text("Selesai!")
        st.balloons()

def download_all_results():
    """Download semua hasil dalam zip"""
    import zipfile
    
    if not os.path.exists(OUTPUT_FOLDER) or not os.listdir(OUTPUT_FOLDER):
        st.warning("Tidak ada hasil yang bisa didownload")
        return
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for filename in os.listdir(OUTPUT_FOLDER):
            filepath = os.path.join(OUTPUT_FOLDER, filename)
            zip_file.write(filepath, filename)
    
    zip_buffer.seek(0)
    
    st.download_button(
        label="📦 Download All Results (ZIP)",
        data=zip_buffer,
        file_name="all_results.zip",
        mime="application/zip"
    )

def render_system_logs():
    """Tampilkan system logs"""
    st.subheader("📊 System Logs & Activity")
    
    # Filter logs
    col_log1, col_log2 = st.columns(2)
    with col_log1:
        log_level = st.selectbox("Filter Level", ["All", "INFO", "SUCCESS", "ERROR", "SYSTEM"], key="log_level")
    with col_log2:
        log_count = st.slider("Jumlah log", 10, 100, 20, key="log_count")
    
    # Display logs
    logs = st.session_state.system_logs[-log_count:] if st.session_state.system_logs else []
    
    if not logs:
        st.info("No logs available")
        return
    
    # Apply filter
    if log_level != "All":
        logs = [log for log in logs if log["type"] == log_level]
    
    # Create log display
    log_container = st.container()
    with log_container:
        for log in reversed(logs):
            log_color = {
                "INFO": "#2196F3",
                "SUCCESS": "#4CAF50",
                "ERROR": "#F44336",
                "SYSTEM": "#FF9800"
            }.get(log["type"], "#9E9E9E")
            
            st.markdown(
                f"""
                <div style='
                    background-color: #f5f5f5;
                    border-left: 4px solid {log_color};
                    padding: 10px;
                    margin-bottom: 8px;
                    border-radius: 0 5px 5px 0;
                '>
                <div style='display: flex; justify-content: space-between;'>
                    <small><strong>{log["type"]}</strong></small>
                    <small>{log["timestamp"]}</small>
                </div>
                <div>{log["message"]}</div>
                </div>
                """,
                unsafe_allow_html=True
            )
    
    # Clear logs button
    if st.button("🗑️ Clear All Logs", type="secondary"):
        st.session_state.system_logs = []
        db.save_data()
        st.rerun()

def render_system_settings():
    """Pengaturan sistem"""
    st.subheader("⚙️ System Configuration")
    
    # Mode settings
    st.markdown("### 🎛️ System Mode")
    current_mode = st.session_state.auto_mode
    
    mode_col1, mode_col2 = st.columns(2)
    with mode_col1:
        new_mode = st.selectbox(
            "Select Mode",
            ["Manual", "Auto"],
            index=1 if current_mode else 0,
            key="system_mode"
        )
    
    with mode_col2:
        if st.button("Apply Mode Change", use_container_width=True):
            st.session_state.auto_mode = (new_mode == "Auto")
            db.add_log(f"Mode changed to {new_mode}", "SYSTEM")
            st.rerun()
    
    # Folder settings
    st.markdown("### 📁 Folder Paths")
    
    col_path1, col_path2, col_path3 = st.columns(3)
    with col_path1:
        st.text_input("Input Folder", INPUT_FOLDER, disabled=True)
    with col_path2:
        st.text_input("Output Folder", OUTPUT_FOLDER, disabled=True)
    with col_path3:
        st.text_input("Archive Folder", ARCHIVE_FOLDER, disabled=True)
    
    # Auto-process settings
    st.markdown("### 🤖 Auto-Process Rules")
    
    col_rule1, col_rule2 = st.columns(2)
    with col_rule1:
        auto_create_task = st.checkbox("Auto-create Kanban tasks", value=True)
        notify_completion = st.checkbox("Notify on completion", value=True)
    with col_rule2:
        keep_archive = st.checkbox("Keep archive files", value=True)
        max_files = st.number_input("Max files in output", 10, 1000, 100)
    
    # Backup/Restore
    st.markdown("### 💾 Backup & Restore")
    
    col_back1, col_back2 = st.columns(2)
    with col_back1:
        if st.button("🔒 Backup Database", use_container_width=True):
            backup_database()
    
    with col_back2:
        backup_file = st.file_uploader("Restore from backup", type=['json'])
        if backup_file and st.button("🔄 Restore Database", use_container_width=True):
            restore_database(backup_file)
    
    # System info
    st.markdown("---")
    st.markdown("### ℹ️ System Information")
    
    info_col1, info_col2 = st.columns(2)
    with info_col1:
        st.write(f"**Python Version:** 3.x")
        st.write(f"**Streamlit Version:** {st.__version__}")
        st.write(f"**Pandas Version:** {pd.__version__}")
    
    with info_col2:
        st.write(f"**Total Tasks:** {sum(len(tasks) for tasks in st.session_state.kanban_tasks.values())}")
        st.write(f"**Total Logs:** {len(st.session_state.system_logs)}")
        st.write(f"**Processed Files:** {len(st.session_state.processed_files)}")

def backup_database():
    """Backup database ke file"""
    backup_data = {
        'kanban_tasks': st.session_state.kanban_tasks,
        'kanban_next_id': st.session_state.kanban_next_id,
        'system_logs': st.session_state.system_logs,
        'processed_files': st.session_state.processed_files,
        'backup_timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    backup_json = json.dumps(backup_data, ensure_ascii=False, indent=2)
    
    st.download_button(
        label="📥 Download Backup",
        data=backup_json,
        file_name=f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        mime="application/json"
    )

def restore_database(backup_file):
    """Restore database dari file"""
    try:
        backup_data = json.load(backup_file)
        
        st.session_state.kanban_tasks = backup_data.get('kanban_tasks', {
            "todo": [], "in_progress": [], "review": [], "done": []
        })
        st.session_state.kanban_next_id = backup_data.get('kanban_next_id', 1)
        st.session_state.system_logs = backup_data.get('system_logs', [])
        st.session_state.processed_files = backup_data.get('processed_files', [])
        
        db.save_data()
        st.success("✅ Database restored successfully!")
        st.rerun()
    except Exception as e:
        st.error(f"❌ Failed to restore: {str(e)}")

# ==================== MAIN APP ====================

def main():
    # Page config
    st.set_page_config(
        page_title="Sistem Otomatis Jadwal Poli",
        page_icon="🏥🤖",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Sidebar
    with st.sidebar:
        st.title("🏥🤖 Auto-System")
        st.markdown("---")
        
        # System status
        status_color = "🟢" if st.session_state.auto_mode else "🔴"
        st.markdown(f"**Status:** {status_color} {'AUTO' if st.session_state.auto_mode else 'MANUAL'}")
        
        # Quick stats
        st.markdown("---")
        st.markdown("**📈 Quick Stats**")
        
        total_tasks = sum(len(tasks) for tasks in st.session_state.kanban_tasks.values())
        pending_tasks = len(st.session_state.kanban_tasks["todo"]) + len(st.session_state.kanban_tasks["in_progress"])
        
        st.metric("Total Tasks", total_tasks)
        st.metric("Pending", pending_tasks)
        
        # Navigation
        st.markdown("---")
        st.markdown("**🔧 Navigation**")
        
        page = st.radio(
            "Go to",
            ["📊 Dashboard", "📋 Kanban", "📁 Files", "📊 Logs", "⚙️ Settings"],
            index=0
        )
        
        st.markdown("---")
        st.caption("v2.0 • Auto-Processing System")
    
    # Main content
    if page == "📊 Dashboard":
        render_dashboard()
    elif page == "📋 Kanban":
        render_kanban_board()
    elif page == "📁 Files":
        render_file_monitor()
    elif page == "📊 Logs":
        render_system_logs()
    elif page == "⚙️ Settings":
        render_system_settings()

# ==================== STARTUP ====================

if __name__ == "__main__":
    # Buat folder jika belum ada
    for folder in [INPUT_FOLDER, OUTPUT_FOLDER, ARCHIVE_FOLDER]:
        os.makedirs(folder, exist_ok=True)
    
    # Jalankan aplikasi
    main()
