from openpyxl import load_workbook


class Validator:
    @staticmethod
    def validate(file):
        try:
            wb = load_workbook(file)
            if "Reguler" not in wb.sheetnames:
                return False, "Sheet 'Reguler' tidak ditemukan"
            if "Poleks" not in wb.sheetnames:
                return False, "Sheet 'Poleks' tidak ditemukan"
            return True, None
        except Exception as e:
            return False, str(e)
