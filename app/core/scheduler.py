import io
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference


class ExcelWriter:

    def __init__(self, config):

        self.config = config
        self.interval = config.interval_minutes
        self.max_e = config.max_poleks_per_slot

        # Warna slot
        self.fill_r = PatternFill("solid", fgColor="C6EFCE")   # Hijau soft
        self.fill_e = PatternFill("solid", fgColor="BDD7EE")   # Biru soft
        self.fill_over = PatternFill("solid", fgColor="FFC7CE")  # Merah soft

        # Warna konflik
        self.fill_conflict = PatternFill("solid", fgColor="FFD966")  # Kuning
        self.fill_conflict_hard = PatternFill("solid", fgColor="FF0000")

        # Header
        self.header_border = Border(
            bottom=Side(border_style="thick")
        )

    # -------------------------------------------------------------------
    # Helper: gabung slot menjadi range waktu
    # -------------------------------------------------------------------
    def _combine_ranges(self, slots):
        if not slots:
            return []

        times = sorted(datetime.strptime(s, "%H:%M") for s in slots)
        out = []
        start = times[0]
        end = start + timedelta(minutes=self.interval)

        for t in times[1:]:
            if t == end:
                end = t + timedelta(minutes=self.interval)
            else:
                out.append((start, end))
                start = t
                end = start + timedelta(minutes=self.interval)

        out.append((start, end))
        return out

    def _format_range(self, a, b):
        return f"{a.strftime('%H:%M')}-{b.strftime('%H:%M')}"

    # -------------------------------------------------------------------
    # 1. Jadwal (utama)
    # -------------------------------------------------------------------
    def write(self, source_file, df_grid, slot_str):

        # df_grid = DF HASIL SCHEDULER (bukan DF upload!)
        wb = load_workbook(source_file)

        if "Jadwal" in wb.sheetnames:
            del wb["Jadwal"]

        ws = wb.create_sheet("Jadwal")

        headers = ["POLI", "JENIS", "HARI", "DOKTER", "JAM"] + slot_str
        ws.append(headers)

        for _, row in df_grid.iterrows():
            ws.append([row.get(h, "") for h in headers])

        # aplikasikan pewarnaan grid
        self._style_jadwal(ws, df_grid, slot_str)

        # Buat semua sheet rekap
        self._sheet_rekap_layanan(wb, df_grid, slot_str)
        self._sheet_rekap_poli(wb, df_grid, slot_str)
        self._sheet_rekap_dokter(wb, df_grid, slot_str)
        self._sheet_peak_hour(wb, df_grid, slot_str)
        self._sheet_conflict_doctor(wb, df_grid, slot_str)
        self._sheet_conflict_map(wb, df_grid, slot_str)
        self._sheet_grafik_poli(wb)

        # Finishing
        self._auto_width(wb)
        self._style_headers(wb)
        self._freeze_headers(wb)

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        buf.seek(0)
        return buf

    # -------------------------------------------------------------------
    # 2. Pewarnaan slot
    # -------------------------------------------------------------------
    def _style_jadwal(self, ws, df, slot_str):

        counter = {hari: {slot: 0 for slot in slot_str} for hari in df["HARI"].unique()}

        for i, rec in enumerate(df.to_dict("records"), start=2):
            hari = rec["HARI"]

            for idx, slot in enumerate(slot_str):
                v = rec.get(slot, "")
                cell = ws.cell(row=i, column=6 + idx)

                if v == "R":
                    cell.fill = self.fill_r

                elif v == "E":
                    counter[hari][slot] += 1
                    if counter[hari][slot] > self.max_e:
                        cell.fill = self.fill_over
                    else:
                        cell.fill = self.fill_e

    # -------------------------------------------------------------------
    # 3. Rekap layanan dokter
    # -------------------------------------------------------------------
    def _sheet_rekap_layanan(self, wb, df, slot_str):

        if "Rekap Layanan" in wb.sheetnames:
            del wb["Rekap Layanan"]

        ws = wb.create_sheet("Rekap Layanan")
        ws.append(["POLI", "HARI", "DOKTER", "JENIS", "WAKTU LAYANAN"])

        for (poli, hari, dokter, jenis), g in df.groupby(["POLI", "HARI", "DOKTER", "JENIS"]):

            active = [s for s in slot_str if g.iloc[0].get(s) in ["R", "E"]]
            merged = self._combine_ranges(active)

            for a, b in merged:
                ws.append([poli, hari, dokter, jenis, self._format_range(a, b)])

    # -------------------------------------------------------------------
    # 4. Rekap Poli
    # -------------------------------------------------------------------
    def _sheet_rekap_poli(self, wb, df, slot_str):

        if "Rekap Poli" in wb.sheetnames:
            del wb["Rekap Poli"]

        ws = wb.create_sheet("Rekap Poli")
        ws.append(["POLI", "HARI", "REGULER (JAM)", "POLEKS (JAM)", "TOTAL"])

        for (poli, hari), g in df.groupby(["POLI", "HARI"]):

            tot_r = sum((g.iloc[0].get(s) == "R") * (self.interval / 60) for s in slot_str)
            tot_e = sum((g.iloc[0].get(s) == "E") * (self.interval / 60) for s in slot_str)

            ws.append([poli, hari, round(tot_r, 2), round(tot_e, 2), round(tot_r + tot_e, 2)])

    # -------------------------------------------------------------------
    # 5. Rekap Dokter
    # -------------------------------------------------------------------
    def _sheet_rekap_dokter(self, wb, df, slot_str):

        if "Rekap Dokter" in wb.sheetnames:
            del wb["Rekap Dokter"]

        ws = wb.create_sheet("Rekap Dokter")
        ws.append(["DOKTER", "HARI", "SHIFT", "TOTAL JAM"])

        for (dokter, hari), g in df.groupby(["DOKTER", "HARI"]):
            active = [s for s in slot_str if g.iloc[0].get(s) in ["R", "E"]]
            merged = self._combine_ranges(active)
            for a, b in merged:
                dur = (b - a).seconds / 3600
                ws.append([dokter, hari, self._format_range(a, b), round(dur, 2)])

    # -------------------------------------------------------------------
    # 6. Peak Hour Analysis
    # -------------------------------------------------------------------
    def _sheet_peak_hour(self, wb, df, slot_str):

        if "Peak Hour Analysis" in wb.sheetnames:
            del wb["Peak Hour Analysis"]

        ws = wb.create_sheet("Peak Hour Analysis")
        ws.append(["HARI", "SLOT", "JUMLAH", "LEVEL"])

        for hari, g in df.groupby("HARI"):
            out = []

            for slot in slot_str:
                vals = (g[slot] == "R") | (g[slot] == "E")
                out.append((slot, vals.sum()))

            max_val = max(v for _, v in out)

            for slot, v in out:
                level = "High" if v == max_val else "Medium" if v >= 5 else "Low"
                ws.append([hari, slot, v, level])

    # -------------------------------------------------------------------
    # 7. Konflik Dokter (tekstual)
    # -------------------------------------------------------------------
    def _sheet_conflict_doctor(self, wb, df, slot_str):

        if "Conflict Dokter" in wb.sheetnames:
            del wb["Conflict Dokter"]

        ws = wb.create_sheet("Conflict Dokter")
        ws.append(["DOKTER", "HARI", "SLOT", "KETERANGAN"])

        for (dokter, hari), g in df.groupby(["DOKTER", "HARI"]):
            for slot in slot_str:
                vals = g[slot].unique()

                if len(vals) > 1 and any(v in ["R", "E"] for v in vals):
                    ws.append([dokter, hari, slot, "Poli berbeda pada waktu sama"])

                if "R" in vals and "E" in vals:
                    ws.append([dokter, hari, slot, "Bentrok Reguler & Poleks"])

    # -------------------------------------------------------------------
    # 8. Peta Konflik Dokter (matrix)
    # -------------------------------------------------------------------
    def _sheet_conflict_map(self, wb, df, slot_str):

        if "Peta Konflik Dokter" in wb.sheetnames:
            del wb["Peta Konflik Dokter"]

        ws = wb.create_sheet("Peta Konflik Dokter")

        doctors = sorted(df["DOKTER"].unique())

        ws.append(["SLOT"] + doctors)

        for slot in slot_str:
            ws.append([slot] + [""] * len(doctors))

        idx_map = {doc: i + 2 for i, doc in enumerate(doctors)}

        for (dokter, hari), g in df.groupby(["DOKTER", "HARI"]):
            col = idx_map[dokter]

            for slot in slot_str:
                vals = g[slot].unique()
                row = 1 + slot_str.index(slot) + 1

                if len(vals) > 1:
                    ws.cell(row, col).fill = self.fill_conflict
                if "R" in vals and "E" in vals:
                    ws.cell(row, col).fill = self.fill_conflict_hard

    # -------------------------------------------------------------------
    # 9. Grafik Beban Poli
    # -------------------------------------------------------------------
    def _sheet_grafik_poli(self, wb):

        if "Grafik Poli" in wb.sheetnames:
            del wb["Grafik Poli"]

        ws = wb.create_sheet("Grafik Poli")

        rp = wb["Rekap Poli"]

        table = {}
        for row in rp.iter_rows(min_row=2, values_only=True):
            poli = row[0]
            tot = row[4]
            table[poli] = table.get(poli, 0) + tot

        ws.append(["POLI", "TOTAL JAM"])
        for p, t in table.items():
            ws.append([p, t])

        chart = BarChart()
        chart.title = "Beban Poli"

        data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data)
        chart.set_categories(cats)

        ws.add_chart(chart, "E5")

    # -------------------------------------------------------------------
    # UTILITIES: styling
    # -------------------------------------------------------------------
    def _auto_width(self, wb):
        for ws in wb.worksheets:
            for col in ws.columns:
                lg = max(len(str(c.value)) if c.value else 0 for c in col)
                col[0].column_letter
                ws.column_dimensions[col[0].column_letter].width = lg + 2

    def _style_headers(self, wb):
        for ws in wb.worksheets:
            for c in ws[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center")
                c.border = self.header_border

    def _freeze_headers(self, wb):
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"

    # -------------------------------------------------------------------
    # TEMPLATE GENERATOR
    # -------------------------------------------------------------------
    def generate_template(self, slot_str):

        wb = Workbook()
        ws = wb.active
        ws.title = "Template"

        headers = ["POLI", "JENIS", "HARI", "DOKTER", "JAM"]
        ws.append(headers)

        ws.append(["Poli Anak", "Reguler", "Senin", "dr. Contoh", "07.00-10.00"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

