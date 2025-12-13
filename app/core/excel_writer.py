"""
ExcelWriter - Modul untuk menulis hasil jadwal ke file Excel
Membuat file Excel dengan multiple sheets: Jadwal, Rekap, Analisis, dll.
"""

import io
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import traceback


class ExcelWriter:
    def __init__(self, config):
        """
        Inisialisasi ExcelWriter dengan konfigurasi
        
        Args:
            config: Config object berisi pengaturan aplikasi
        """
        self.config = config
        self.interval = config.interval_minutes
        self.max_e = config.max_poleks_per_slot
        
        print(f"✅ ExcelWriter initialized with config:")
        print(f"   - Interval: {self.interval} minutes")
        print(f"   - Max poleks per slot: {self.max_e}")
        print(f"   - Color rules: R=Green, E(≤{self.max_e})=Blue, E(>{self.max_e})=Red")
        
        # ======================================================
        # DEFINE COLORS - PERBAIKAN WARNA
        # ======================================================
        
        # WARNA UTAMA UNTUK SLOT
        self.fill_r = PatternFill("solid", fgColor="92D050")     # HIJAU TERANG - Reguler
        self.fill_e = PatternFill("solid", fgColor="4BACC6")     # BIRU - Poleks (dalam batas)
        self.fill_over = PatternFill("solid", fgColor="FF0000")  # MERAH - Poleks (overload)
        
        # WARNA TAMBAHAN
        self.fill_header = PatternFill("solid", fgColor="366092")  # Biru tua - Header
        self.fill_gray = PatternFill("solid", fgColor="F2F2F2")    # Abu-abu - Alternating rows
        self.fill_total = PatternFill("solid", fgColor="FFE699")   # Kuning muda - Total
        
        # Warna konflik
        self.fill_conflict = PatternFill("solid", fgColor="FFD966")  # Kuning - Konflik ringan
        self.fill_conflict_hard = PatternFill("solid", fgColor="FF4500")  # Merah tua - Konflik berat
        
        # Warna untuk Kanban
        self.fill_kanban_header = PatternFill("solid", fgColor="7030A0")  # Ungu - Header Kanban
        self.fill_high = PatternFill("solid", fgColor="FF4D4F")          # Merah - Prioritas Tinggi
        self.fill_medium = PatternFill("solid", fgColor="FAAD14")        # Kuning - Prioritas Sedang
        self.fill_low = PatternFill("solid", fgColor="52C41A")           # Hijau - Prioritas Rendah
        
        # Warna label Kanban
        self.fill_overload = PatternFill("solid", fgColor="FF7875")      # Merah muda - Overload
        self.fill_konflik = PatternFill("solid", fgColor="FF9C6E")       # Oranye - Konflik
        self.fill_kosong = PatternFill("solid", fgColor="69C0FF")        # Biru muda - Kosong
        self.fill_distribusi = PatternFill("solid", fgColor="95DE64")    # Hijau muda - Distribusi
        self.fill_beban = PatternFill("solid", fgColor="B37FEB")         # Ungu muda - Beban
        self.fill_review = PatternFill("solid", fgColor="FFD666")        # Kuning muda - Review
        self.fill_optimal = PatternFill("solid", fgColor="5CDBD3")       # Cyan - Optimal
        
        # Font
        self.font_header = Font(bold=True, color="FFFFFF", size=11)
        self.font_normal = Font(size=10)
        self.font_bold = Font(bold=True)
        self.font_small = Font(size=9)
        
        # Alignment
        self.align_center = Alignment(horizontal="center", vertical="center")
        self.align_left = Alignment(horizontal="left", vertical="center")
        self.align_right = Alignment(horizontal="right", vertical="center")
        
        # Borders
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        self.thick_border = Border(
            bottom=Side(style="thick")
        )
        
        print("✅ ExcelWriter styling configured")
    
    # ======================================================
    # MAIN WRITE METHOD
    # ======================================================
    
    def write(self, source_file, df_grid, slot_str, kanban_data=None):
        """
        Tulis hasil jadwal ke file Excel dengan multiple sheets
        
        Args:
            source_file: File Excel asli (BytesIO atau path) sebagai template
            df_grid: DataFrame hasil scheduler (format grid)
            slot_str: List string slot waktu
            kanban_data: Data kanban board (default: None)
            
        Returns:
            BytesIO buffer berisi file Excel
        """
        print(f"📝 ExcelWriter.write() called")
        print(f"   - df_grid shape: {df_grid.shape if df_grid is not None else 'None'}")
        print(f"   - slot_str length: {len(slot_str) if slot_str else 0}")
        print(f"   - Max poleks per slot: {self.max_e}")
        print(f"   - Kanban data: {'Available' if kanban_data else 'Not available'}")
        
        try:
            # Load atau buat workbook
            wb = self._load_or_create_workbook(source_file)
            
            print(f"✅ Workbook ready with {len(wb.sheetnames)} sheets")
            
            # Debug: tampilkan distribusi poleks
            if df_grid is not None and not df_grid.empty:
                self._debug_poleks_distribution(df_grid, slot_str)
            
            # Urutan pembuatan sheets
            sheets_to_create = [
                ("Summary", lambda wb, df, slots: self._create_summary_sheet(wb, df, slots)),
                ("Jadwal", self._create_jadwal_sheet),
                ("Rekap Layanan", lambda wb, df, slots: self._create_rekap_layanan_sheet(wb, df, slots)),
                ("Rekap Poli", lambda wb, df, slots: self._create_rekap_poli_sheet(wb, df, slots)),
                ("Rekap Dokter", lambda wb, df, slots: self._create_rekap_dokter_sheet(wb, df, slots)),
                ("Peak Hour Analysis", lambda wb, df, slots: self._create_peak_hour_sheet(wb, df, slots)),
                ("Conflict Dokter", lambda wb, df, slots: self._create_conflict_doctor_sheet(wb, df, slots)),
                ("Peta Konflik Dokter", lambda wb, df, slots: self._create_conflict_map_sheet(wb, df, slots)),
                ("Grafik Poli", lambda wb, df, slots: self._create_grafik_poli_sheet(wb, df)),
                ("Kanban Board", lambda wb, df, slots: self._create_kanban_sheet(wb, df, slots, kanban_data)),
            ]
            
            # Buat semua sheets
            for sheet_name, create_func in sheets_to_create:
                print(f"Creating '{sheet_name}' sheet...")
                try:
                    create_func(wb, df_grid, slot_str)
                except Exception as e:
                    print(f"⚠️ Error creating sheet '{sheet_name}': {e}")
                    print(traceback.format_exc())
            
            # Apply styling ke semua sheets
            print("Applying styling to all sheets...")
            self._apply_styling_to_all_sheets(wb)
            
            # Auto adjust column widths
            print("Auto-adjusting column widths...")
            self._auto_adjust_column_widths(wb)
            
            # Reorder sheets untuk UX yang lebih baik
            self._reorder_sheets(wb)
            
            # Save to buffer
            print("Saving to buffer...")
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            
            file_size = buf.getbuffer().nbytes
            print(f"✅ Excel file created successfully: {file_size:,} bytes")
            return buf
            
        except Exception as e:
            print(f"❌ Error in ExcelWriter.write(): {e}")
            print(traceback.format_exc())
            # Fallback: buat workbook minimal
            return self._create_fallback_workbook(df_grid, slot_str)
    
    def _debug_poleks_distribution(self, df_grid, slot_str):
        """Debug: tampilkan distribusi poleks per slot"""
        print("\n📊 DEBUG: POLEKS DISTRIBUTION PER HARI & SLOT")
        print("=" * 60)
        
        if df_grid is None or df_grid.empty:
            print("No data available")
            return
        
        # Group by hari
        for hari in sorted(df_grid["HARI"].unique()):
            hari_data = df_grid[df_grid["HARI"] == hari]
            print(f"\n📅 HARI: {hari}")
            print("-" * 40)
            
            overload_slots = []
            
            for slot in slot_str[:15]:  # Tampilkan 15 slot pertama
                if slot in hari_data.columns:
                    e_count = (hari_data[slot] == "E").sum()
                    r_count = (hari_data[slot] == "R").sum()
                    
                    if e_count > 0 or r_count > 0:
                        status = ""
                        if e_count > 0:
                            if e_count <= self.max_e:
                                status = f"OK ({e_count} Poleks)"
                            else:
                                status = f"OVERLOAD! {e_count} > {self.max_e}"
                                overload_slots.append((slot, e_count))
                        
                        print(f"  {slot}: {r_count}R {e_count}E - {status}")
            
            # Tampilkan warning jika ada overload
            if overload_slots:
                print(f"\n  ⚠️ PERINGATAN OVERLOAD di {hari}:")
                for slot, count in overload_slots:
                    print(f"    - {slot}: {count} Poleks (batas: {self.max_e})")
        
        print("\n" + "=" * 60)
    
    def _load_or_create_workbook(self, source_file):
        """Load workbook dari source atau buat baru"""
        try:
            if hasattr(source_file, 'read'):
                # BytesIO atau file-like object
                source_file.seek(0)
                wb = load_workbook(source_file)
                
                # Hapus sheet yang mungkin mengganggu
                sheets_to_remove = ["Jadwal", "Rekap Layanan", "Rekap Poli", 
                                  "Rekap Dokter", "Peak Hour Analysis", 
                                  "Conflict Dokter", "Peta Konflik Dokter",
                                  "Grafik Poli", "Summary", "Kanban Board"]
                
                for sheet in sheets_to_remove:
                    if sheet in wb.sheetnames:
                        del wb[sheet]
                
                return wb
                
        except Exception as e:
            print(f"⚠️ Could not load source workbook: {e}")
        
        # Buat workbook baru
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        return wb
    
    def _create_fallback_workbook(self, df_grid, slot_str):
        """Buat workbook fallback jika error"""
        print("Creating fallback workbook...")
        
        try:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            
            # Buat sheet Jadwal saja
            ws = wb.create_sheet("Jadwal")
            
            if df_grid is not None and not df_grid.empty:
                # Header sederhana
                headers = ["POLI", "JENIS", "HARI", "DOKTER", "JAM"] + slot_str
                ws.append(headers)
                
                # Data
                for _, row in df_grid.iterrows():
                    ws.append([row.get(h, "") for h in headers])
                
                # Apply basic styling
                self._style_jadwal_sheet_fallback(ws, df_grid, slot_str)
            
            # Simpan ke buffer
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            
            return buf
            
        except Exception as e:
            print(f"❌ Even fallback failed: {e}")
            # Buat workbook kosong
            wb = Workbook()
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf
    
    # ======================================================
    # SHEET CREATION METHODS
    # ======================================================
    
    def _create_jadwal_sheet(self, wb, df_grid, slot_str):
        """Buat sheet Jadwal utama"""
        # Hapus sheet lama jika ada
        if "Jadwal" in wb.sheetnames:
            ws_index = wb.sheetnames.index("Jadwal")
            del wb[wb.sheetnames[ws_index]]
        
        ws = wb.create_sheet("Jadwal")  # Buat sebagai sheet pertama
        
        # Header
        headers = ["POLI", "JENIS", "HARI", "DOKTER", "JAM"] + slot_str
        ws.append(headers)
        
        # Data
        if df_grid is not None and not df_grid.empty:
            for _, row in df_grid.iterrows():
                ws.append([row.get(h, "") for h in headers])
        else:
            # Tambahkan contoh jika tidak ada data
            ws.append(["Poli Anak", "Reguler", "Senin", "dr. Contoh", "07:30-10:00"] + [""] * len(slot_str))
        
        # Apply styling DENGAN PERBAIKAN WARNA
        self._style_jadwal_sheet(ws, df_grid, slot_str)
        
        # Freeze header row dan beberapa kolom
        ws.freeze_panes = "F2"
        
        return ws
    
    def _style_jadwal_sheet(self, ws, df_grid, slot_str):
        """Style sheet Jadwal dengan warna sesuai aturan"""
        if ws.max_row <= 1:
            return
        
        print(f"   Styling Jadwal sheet: {ws.max_row-1} rows, {ws.max_column} columns")
        
        # 1. STYLE HEADER
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.fill_header
            cell.font = self.font_header
            cell.alignment = self.align_center
            cell.border = self.thin_border
        
        # 2. STYLE DATA ROWS (kolom metadata)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                
                # Alignment berdasarkan kolom
                if col <= 4:  # POLI to HARI
                    cell.alignment = self.align_left
                elif col == 5:  # JAM
                    cell.alignment = self.align_center
                else:  # Slot waktu
                    cell.alignment = self.align_center
            
            # Alternate row background untuk kolom metadata saja
            if row % 2 == 0:
                for col in range(1, 6):  # Hanya kolom metadata (A sampai E)
                    ws.cell(row=row, column=col).fill = self.fill_gray
        
        # 3. WARNA SLOT WAKTU BERDASARKAN ATURAN
        if df_grid is not None and not df_grid.empty and ws.max_row > 1:
            print(f"   Applying color rules to time slots...")
            
            # Step 1: Kumpulkan semua baris E per (hari, slot) untuk menentukan urutan
            poleks_tracking = {}
            
            for row_idx in range(2, min(len(df_grid) + 2, ws.max_row + 1)):
                hari_cell = ws.cell(row=row_idx, column=3)  # Kolom C = HARI
                hari = str(hari_cell.value).strip() if hari_cell.value else ""
                
                if not hari:
                    continue
                
                for col_idx, slot in enumerate(slot_str, start=6):  # Kolom F dst
                    if col_idx <= ws.max_column:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = str(cell.value).strip() if cell.value else ""
                        
                        if value == "E":
                            key = (hari, slot)
                            if key not in poleks_tracking:
                                poleks_tracking[key] = []
                            poleks_tracking[key].append(row_idx)
            
            # Step 2: Apply warna berdasarkan aturan
            overload_count = 0
            
            for row_idx in range(2, min(len(df_grid) + 2, ws.max_row + 1)):
                hari_cell = ws.cell(row=row_idx, column=3)
                hari = str(hari_cell.value).strip() if hari_cell.value else ""
                
                if not hari:
                    continue
                
                for col_idx, slot in enumerate(slot_str, start=6):
                    if col_idx <= ws.max_column:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = str(cell.value).strip() if cell.value else ""
                        
                        # Reset fill ke default (transparan)
                        cell.fill = PatternFill(fill_type=None)
                        
                        if value == "R":
                            # REGULER: HIJAU
                            cell.fill = self.fill_r
                            cell.value = "R"
                        
                        elif value == "E":
                            key = (hari, slot)
                            
                            if key in poleks_tracking:
                                e_rows = poleks_tracking[key]
                                
                                # Cari posisi baris ini dalam daftar E untuk slot ini
                                if row_idx in e_rows:
                                    position = e_rows.index(row_idx)
                                    
                                    if position < self.max_e:
                                        # Dalam batas: BIRU
                                        cell.fill = self.fill_e
                                    else:
                                        # Melebihi batas: MERAH
                                        cell.fill = self.fill_over
                                        overload_count += 1
                                else:
                                    # Tidak ditemukan: BIRU (fallback)
                                    cell.fill = self.fill_e
                            else:
                                # Tidak ada tracking: BIRU (fallback)
                                cell.fill = self.fill_e
                        
                        else:
                            # Kosong: hapus nilai dan biarkan tanpa warna
                            cell.value = ""
                        
                        # Selalu center alignment
                        cell.alignment = self.align_center
            
            # Debug info
            if overload_count > 0:
                print(f"   ⚠️ Found {overload_count} overloaded Poleks slots (colored RED)")
        
        print(f"   ✅ Jadwal sheet styling completed")
    
    def _style_jadwal_sheet_fallback(self, ws, df_grid, slot_str):
        """Style fallback sederhana"""
        if ws.max_row <= 1:
            return
        
        # Header
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.fill_header
            cell.font = self.font_header
            cell.alignment = self.align_center
        
        # Basic data styling
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                
                if col <= 4:
                    cell.alignment = self.align_left
                else:
                    cell.alignment = self.align_center
                
                # Simple coloring for R and E
                if col >= 6 and cell.value:
                    value = str(cell.value).strip()
                    if value == "R":
                        cell.fill = self.fill_r
                    elif value == "E":
                        cell.fill = self.fill_e
    
    def _create_rekap_layanan_sheet(self, wb, df_grid, slot_str):
        """Buat sheet Rekap Layanan"""
        if "Rekap Layanan" in wb.sheetnames:
            del wb["Rekap Layanan"]
        
        ws = wb.create_sheet("Rekap Layanan")
        ws.append(["POLI", "HARI", "DOKTER", "JENIS", "WAKTU LAYANAN"])
        
        if df_grid is not None and not df_grid.empty:
            processed_combinations = set()
            
            for _, row in df_grid.iterrows():
                poli = row["POLI"]
                hari = row["HARI"]
                dokter = row["DOKTER"]
                jenis = row["JENIS"]
                
                combination_key = f"{poli}|{hari}|{dokter}|{jenis}"
                if combination_key in processed_combinations:
                    continue
                
                processed_combinations.add(combination_key)
                
                # Kumpulkan slot aktif
                active_slots = []
                for slot in slot_str:
                    if slot in row and pd.notna(row[slot]) and row[slot] in ["R", "E"]:
                        active_slots.append(slot)
                
                if active_slots:
                    time_ranges = self._combine_slots_to_ranges(active_slots, slot_str)
                    
                    for time_range in time_ranges:
                        ws.append([poli, hari, dokter, jenis, time_range])
        
        # Style
        self._style_rekap_sheet(ws)
        ws.freeze_panes = "A2"
        
        return ws
    
    def _create_rekap_poli_sheet(self, wb, df_grid, slot_str):
        """Buat sheet Rekap Poli"""
        if "Rekap Poli" in wb.sheetnames:
            del wb["Rekap Poli"]
        
        ws = wb.create_sheet("Rekap Poli")
        ws.append(["POLI", "HARI", "REGULER (JAM)", "POLEKS (JAM)", "TOTAL JAM"])
        
        if df_grid is not None and not df_grid.empty:
            # Group by poli dan hari
            poli_stats = {}
            
            for _, row in df_grid.iterrows():
                poli = row["POLI"]
                hari = row["HARI"]
                key = (poli, hari)
                
                if key not in poli_stats:
                    poli_stats[key] = {"R": 0, "E": 0}
                
                # Hitung slot R dan E
                for slot in slot_str:
                    if slot in row and pd.notna(row[slot]):
                        if row[slot] == "R":
                            poli_stats[key]["R"] += 1
                        elif row[slot] == "E":
                            poli_stats[key]["E"] += 1
            
            # Tulis data
            for (poli, hari), counts in poli_stats.items():
                hours_r = round(counts["R"] * self.interval / 60, 2)
                hours_e = round(counts["E"] * self.interval / 60, 2)
                total_hours = hours_r + hours_e
                
                ws.append([poli, hari, hours_r, hours_e, total_hours])
        
        # Add totals row
        if ws.max_row > 1:
            last_row = ws.max_row
            ws.append(["TOTAL", "", 
                      f"=SUM(C2:C{last_row})", 
                      f"=SUM(D2:D{last_row})", 
                      f"=SUM(E2:E{last_row})"])
            
            # Style total row
            total_row = ws.max_row
            for col in range(1, 6):
                cell = ws.cell(row=total_row, column=col)
                cell.fill = self.fill_total
                cell.font = self.font_bold
        
        # Style
        self._style_rekap_sheet(ws)
        ws.freeze_panes = "A2"
        
        return ws
    
    def _create_rekap_dokter_sheet(self, wb, df_grid, slot_str):
        """Buat sheet Rekap Dokter"""
        if "Rekap Dokter" in wb.sheetnames:
            del wb["Rekap Dokter"]
        
        ws = wb.create_sheet("Rekap Dokter")
        ws.append(["DOKTER", "HARI", "SHIFT", "TOTAL JAM"])
        
        if df_grid is not None and not df_grid.empty:
            # Group by dokter dan hari
            doctor_shifts = {}
            
            for _, row in df_grid.iterrows():
                dokter = row["DOKTER"]
                hari = row["HARI"]
                key = (dokter, hari)
                
                if key not in doctor_shifts:
                    doctor_shifts[key] = []
                
                # Kumpulkan slot aktif
                active_slots = []
                for slot in slot_str:
                    if slot in row and pd.notna(row[slot]) and row[slot] in ["R", "E"]:
                        active_slots.append(slot)
                
                if active_slots:
                    doctor_shifts[key].extend(active_slots)
            
            # Tulis data
            for (dokter, hari), slots in doctor_shifts.items():
                if slots:
                    # Gabungkan slot berurutan
                    unique_slots = sorted(set(slots))
                    time_ranges = self._combine_slots_to_ranges(unique_slots, slot_str)
                    
                    for time_range in time_ranges:
                        # Hitung durasi
                        duration = self._calculate_duration(time_range, slot_str)
                        ws.append([dokter, hari, time_range, round(duration, 2)])
        
        # Style
        self._style_rekap_sheet(ws)
        ws.freeze_panes = "A2"
        
        return ws
    
    def _create_peak_hour_sheet(self, wb, df_grid, slot_str):
        """Buat sheet Peak Hour Analysis"""
        if "Peak Hour Analysis" in wb.sheetnames:
            del wb["Peak Hour Analysis"]
        
        ws = wb.create_sheet("Peak Hour Analysis")
        ws.append(["HARI", "SLOT", "JUMLAH DOKTER", "LEVEL"])
        
        if df_grid is not None and not df_grid.empty:
            # Hitung per hari
            for hari in sorted(df_grid["HARI"].unique()):
                hari_data = df_grid[df_grid["HARI"] == hari]
                
                for slot in slot_str:
                    if slot in hari_data.columns:
                        count = ((hari_data[slot] == "R") | (hari_data[slot] == "E")).sum()
                        
                        if count > 0:
                            # Tentukan level
                            if count >= 10:
                                level = "VERY HIGH"
                            elif count >= 7:
                                level = "HIGH"
                            elif count >= 4:
                                level = "MEDIUM"
                            else:
                                level = "LOW"
                            
                            ws.append([hari, slot, count, level])
        
        # Style
        self._style_rekap_sheet(ws)
        ws.freeze_panes = "A2"
        
        return ws
    
    def _create_conflict_doctor_sheet(self, wb, df_grid, slot_str):
        """Buat sheet Conflict Dokter"""
        if "Conflict Dokter" in wb.sheetnames:
            del wb["Conflict Dokter"]
        
        ws = wb.create_sheet("Conflict Dokter")
        ws.append(["DOKTER", "HARI", "SLOT", "KETERANGAN", "TINGKAT"])
        
        if df_grid is not None and not df_grid.empty:
            conflicts = self._find_doctor_conflicts(df_grid, slot_str)
            
            for conflict in conflicts:
                ws.append([
                    conflict["dokter"],
                    conflict["hari"],
                    conflict["slot"],
                    conflict["keterangan"],
                    conflict["tingkat"]
                ])
        
        # Jika tidak ada konflik
        if ws.max_row == 1:
            ws.append(["", "", "", "✅ Tidak ada konflik ditemukan", "INFO"])
        
        # Style
        self._style_conflict_sheet(ws)
        ws.freeze_panes = "A2"
        
        return ws
    
    def _create_conflict_map_sheet(self, wb, df_grid, slot_str):
        """Buat sheet Peta Konflik Dokter"""
        if "Peta Konflik Dokter" in wb.sheetnames:
            del wb["Peta Konflik Dokter"]
        
        ws = wb.create_sheet("Peta Konflik Dokter")
        
        # Get unique doctors
        doctors = sorted(df_grid["DOKTER"].unique()) if df_grid is not None and not df_grid.empty else []
        
        if not doctors:
            ws.append(["Tidak ada data dokter"])
            return ws
        
        # Header row
        header = ["SLOT"] + doctors
        ws.append(header)
        
        # Data rows
        for slot in slot_str:
            row = [slot]
            for _ in doctors:
                row.append("")  # Placeholder
            ws.append(row)
        
        # Fill conflict data
        if df_grid is not None and not df_grid.empty:
            for (dokter, hari), group in df_grid.groupby(["DOKTER", "HARI"]):
                if dokter in doctors:
                    col_idx = doctors.index(dokter) + 2  # +1 untuk header, +1 untuk 1-based indexing
                    
                    for slot in slot_str:
                        if slot in group.columns:
                            row_idx = slot_str.index(slot) + 2  # +1 untuk header, +1 untuk 1-based indexing
                            cell = ws.cell(row=row_idx, column=col_idx)
                            
                            values = group[slot].unique()
                            
                            if len(values) > 1 and any(v in ["R", "E"] for v in values):
                                cell.value = "⚠️"
                                cell.fill = self.fill_conflict
                                cell.alignment = self.align_center
                            
                            if "R" in values and "E" in values:
                                cell.value = "🚨"
                                cell.fill = self.fill_conflict_hard
                                cell.alignment = self.align_center
        
        # Style
        self._style_conflict_map_sheet(ws, len(doctors))
        
        return ws
    
    def _create_grafik_poli_sheet(self, wb, df_grid):
        """Buat sheet Grafik Poli"""
        if "Grafik Poli" in wb.sheetnames:
            del wb["Grafik Poli"]
        
        ws = wb.create_sheet("Grafik Poli")
        
        # Coba ambil data dari Rekap Poli
        chart_data = []
        
        if "Rekap Poli" in wb.sheetnames:
            try:
                rp_ws = wb["Rekap Poli"]
                
                # Kumpulkan total per poli (abaikan baris TOTAL)
                poli_totals = {}
                for row in rp_ws.iter_rows(min_row=2, max_col=5, values_only=True):
                    if row and row[0] and row[0] != "TOTAL" and row[4] is not None:
                        poli = str(row[0]).strip()
                        total = float(row[4]) if isinstance(row[4], (int, float)) else 0
                        poli_totals[poli] = poli_totals.get(poli, 0) + total
                
                # Konversi ke list untuk chart
                for poli, total in sorted(poli_totals.items(), key=lambda x: x[1], reverse=True):
                    chart_data.append([poli, total])
                
            except Exception as e:
                print(f"⚠️ Error extracting chart data: {e}")
        
        # Jika tidak ada data, buat dummy
        if not chart_data:
            chart_data = [
                ["Poli Anak", 24.5],
                ["Poli Dalam", 18.2],
                ["Poli Bedah", 15.8],
                ["Poli Jantung", 12.3],
                ["Poli Lainnya", 8.7]
            ]
        
        # Write data
        ws.append(["POLI", "TOTAL JAM"])
        for data in chart_data:
            ws.append(data)
        
        # Create chart jika ada data
        if len(chart_data) > 0:
            try:
                chart = BarChart()
                chart.title = "Beban Poli (Total Jam)"
                chart.style = 10
                chart.y_axis.title = "Total Jam"
                chart.x_axis.title = "Poli"
                chart.height = 15
                chart.width = 25
                
                data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
                categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                
                # Tambahkan chart ke sheet
                ws.add_chart(chart, "E5")
                
            except Exception as e:
                print(f"⚠️ Could not create chart: {e}")
        
        # Style
        self._style_chart_sheet(ws)
        
        return ws
    
def _create_kanban_sheet(self, wb, df_grid, slot_str, kanban_data=None):
    """Buat sheet Kanban Board sebagai laporan text sederhana"""
    if "Kanban Board" in wb.sheetnames:
        del wb["Kanban Board"]
    
    ws = wb.create_sheet("Kanban Board")
    
    # Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "📌 LAPORAN KANBAN BOARD - MANAJEMEN JADWAL DOKTER"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="7030A0")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Timestamp
    ws.append([])  # Empty row
    ws.append(["Dibuat pada", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Lokasi File", "Aplikasi Jadwal Dokter > Tab Kanban"])
    ws.append([])
    
    # Jika tidak ada kanban_data, buat default
    if not kanban_data:
        kanban_data = self._generate_kanban_from_analysis(df_grid, slot_str)
    
    # Define kanban columns order
    kanban_columns_order = [
        "⚠️ MASALAH JADWAL",
        "🔧 PERLU PENYESUAIAN", 
        "⏳ DALAM PROSES",
        "✅ OPTIMAL"
    ]
    
    row = 5  # Start row for content
    
    # Write each kanban column
    for column_name in kanban_columns_order:
        # Column header
        ws.merge_cells(f"A{row}:F{row}")
        header_cell = ws[f"A{row}"]
        header_cell.value = column_name
        header_cell.font = Font(bold=True, size=12, color="FFFFFF")
        header_cell.fill = PatternFill("solid", fgColor="366092")
        header_cell.alignment = Alignment(horizontal="left", vertical="center")
        header_cell.border = self.thin_border
        
        # Get cards for this column
        cards = kanban_data.get(column_name, [])
        card_count = len(cards)
        
        row += 1
        
        # If no cards
        if card_count == 0:
            ws.merge_cells(f"A{row}:F{row}")
            empty_cell = ws[f"A{row}"]
            empty_cell.value = "   Tidak ada kartu"
            empty_cell.font = Font(italic=True, color="666666")
            empty_cell.alignment = Alignment(horizontal="left", vertical="center")
            row += 2
            continue
        
        # Write each card
        for i, card in enumerate(cards, 1):
            # Card text
            text = card.get("text", "")
            label = card.get("label", "")
            priority = card.get("priority", "Medium")
            
            # Format card display
            card_display = []
            card_display.append(f"{i}. {text}")
            card_display.append(f"   • Label: {label}")
            card_display.append(f"   • Prioritas: {priority}")
            
            # Add any data details if available
            if "data" in card:
                data = card["data"]
                if "type" in data:
                    if data["type"] == "overload":
                        card_display.append(f"   • Detail: {data['hari']} {data['slot']}, {data['count']} Poleks (batas {data['max']})")
                    elif data["type"] == "conflict":
                        card_display.append(f"   • Detail: {data['dokter']} - {data['hari']}, {len(data['conflicts'])} konflik")
                    elif data["type"] == "empty":
                        card_display.append(f"   • Detail: {data['poli']} - {data['hari']}, {data['empty_slots']} slot kosong")
                    elif data["type"] == "distribution":
                        card_display.append(f"   • Detail: {data['poli']}, {data['morning_pct']:.0f}% pagi, {data['afternoon_pct']:.0f}% sore")
                    elif data["type"] == "optimal":
                        card_display.append(f"   • Detail: {data['hari']} {data['slot']}, {data['doctor_count']} dokter")
            
            # Write each line of the card
            for line in card_display:
                ws.merge_cells(f"A{row}:F{row}")
                cell = ws[f"A{row}"]
                cell.value = line
                
                # Style based on priority
                if priority == "High":
                    cell.font = Font(bold=True, color="FF0000")
                elif priority == "Medium":
                    cell.font = Font(color="FF9900")
                else:  # Low
                    cell.font = Font(color="00AA00")
                
                if line.startswith("   •"):
                    cell.font = Font(size=9, color="666666")
                
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                row += 1
            
            # Add small gap between cards
            row += 1
        
        # Add gap between columns
        row += 1
    
    # Statistics section
    stats_row = row + 2
    ws.merge_cells(f"A{stats_row}:F{stats_row}")
    stats_title = ws[f"A{stats_row}"]
    stats_title.value = "📊 STATISTIK KANBAN"
    stats_title.font = Font(bold=True, size=12)
    stats_title.alignment = Alignment(horizontal="center", vertical="center")
    
    # Calculate statistics
    total_cards = 0
    high_priority = 0
    medium_priority = 0
    low_priority = 0
    
    for column_name, cards in kanban_data.items():
        total_cards += len(cards)
        for card in cards:
            priority = card.get("priority", "Medium")
            if priority == "High":
                high_priority += 1
            elif priority == "Medium":
                medium_priority += 1
            else:
                low_priority += 1
    
    # Write statistics in table format
    stats_row += 2
    
    # Table header
    ws.merge_cells(f"A{stats_row}:B{stats_row}")
    header1 = ws[f"A{stats_row}"]
    header1.value = "METRIK"
    header1.font = Font(bold=True)
    header1.fill = PatternFill("solid", fgColor="D9E1F2")
    header1.alignment = Alignment(horizontal="center", vertical="center")
    header1.border = self.thin_border
    
    ws.merge_cells(f"C{stats_row}:D{stats_row}")
    header2 = ws[f"C{stats_row}"]
    header2.value = "JUMLAH"
    header2.font = Font(bold=True)
    header2.fill = PatternFill("solid", fgColor="D9E1F2")
    header2.alignment = Alignment(horizontal="center", vertical="center")
    header2.border = self.thin_border
    
    ws.merge_cells(f"E{stats_row}:F{stats_row}")
    header3 = ws[f"E{stats_row}"]
    header3.value = "PERSENTASE"
    header3.font = Font(bold=True)
    header3.fill = PatternFill("solid", fgColor="D9E1F2")
    header3.alignment = Alignment(horizontal="center", vertical="center")
    header3.border = self.thin_border
    
    # Data rows
    stats_data = [
        ["Total Kartu", total_cards, f"{100:.1f}%" if total_cards > 0 else "0%"],
        ["Prioritas Tinggi", high_priority, f"{(high_priority/total_cards*100):.1f}%" if total_cards > 0 else "0%"],
        ["Prioritas Sedang", medium_priority, f"{(medium_priority/total_cards*100):.1f}%" if total_cards > 0 else "0%"],
        ["Prioritas Rendah", low_priority, f"{(low_priority/total_cards*100):.1f}%" if total_cards > 0 else "0%"],
    ]
    
    for i, (label, value, pct) in enumerate(stats_data):
        row_num = stats_row + 1 + i
        
        # Metric
        ws.merge_cells(f"A{row_num}:B{row_num}")
        cell1 = ws[f"A{row_num}"]
        cell1.value = label
        cell1.alignment = Alignment(horizontal="left", vertical="center")
        cell1.border = self.thin_border
        
        # Count
        ws.merge_cells(f"C{row_num}:D{row_num}")
        cell2 = ws[f"C{row_num}"]
        cell2.value = value
        cell2.alignment = Alignment(horizontal="center", vertical="center")
        cell2.border = self.thin_border
        
        # Percentage
        ws.merge_cells(f"E{row_num}:F{row_num}")
        cell3 = ws[f"E{row_num}"]
        cell3.value = pct
        cell3.alignment = Alignment(horizontal="center", vertical="center")
        cell3.border = self.thin_border
        
        # Highlight important rows
        if label == "Total Kartu":
            cell1.font = Font(bold=True)
            cell2.font = Font(bold=True)
            cell3.font = Font(bold=True)
    
    # Column distribution
    dist_row = stats_row + len(stats_data) + 3
    
    ws.merge_cells(f"A{dist_row}:F{dist_row}")
    dist_title = ws[f"A{dist_row}"]
    dist_title.value = "📋 DISTRIBUSI PER KOLOM"
    dist_title.font = Font(bold=True, size=11)
    dist_title.alignment = Alignment(horizontal="center", vertical="center")
    
    dist_row += 1
    
    # Table header for column distribution
    ws.merge_cells(f"A{dist_row}:B{dist_row}")
    dheader1 = ws[f"A{dist_row}"]
    dheader1.value = "KOLOM"
    dheader1.font = Font(bold=True)
    dheader1.fill = PatternFill("solid", fgColor="F2F2F2")
    dheader1.alignment = Alignment(horizontal="left", vertical="center")
    dheader1.border = self.thin_border
    
    ws.merge_cells(f"C{dist_row}:D{dist_row}")
    dheader2 = ws[f"C{dist_row}"]
    dheader2.value = "JUMLAH KARTU"
    dheader2.font = Font(bold=True)
    dheader2.fill = PatternFill("solid", fgColor="F2F2F2")
    dheader2.alignment = Alignment(horizontal="center", vertical="center")
    dheader2.border = self.thin_border
    
    ws.merge_cells(f"E{dist_row}:F{dist_row}")
    dheader3 = ws[f"E{dist_row}"]
    dheader3.value = "% DARI TOTAL"
    dheader3.font = Font(bold=True)
    dheader3.fill = PatternFill("solid", fgColor="F2F2F2")
    dheader3.alignment = Alignment(horizontal="center", vertical="center")
    dheader3.border = self.thin_border
    
    # Column distribution data
    for i, column_name in enumerate(kanban_columns_order):
        row_num = dist_row + 1 + i
        cards = kanban_data.get(column_name, [])
        count = len(cards)
        pct = f"{(count/total_cards*100):.1f}%" if total_cards > 0 else "0%"
        
        # Column name
        ws.merge_cells(f"A{row_num}:B{row_num}")
        cell1 = ws[f"A{row_num}"]
        cell1.value = column_name
        cell1.alignment = Alignment(horizontal="left", vertical="center")
        cell1.border = self.thin_border
        
        # Count
        ws.merge_cells(f"C{row_num}:D{row_num}")
        cell2 = ws[f"C{row_num}"]
        cell2.value = count
        cell2.alignment = Alignment(horizontal="center", vertical="center")
        cell2.border = self.thin_border
        
        # Percentage
        ws.merge_cells(f"E{row_num}:F{row_num}")
        cell3 = ws[f"E{row_num}"]
        cell3.value = pct
        cell3.alignment = Alignment(horizontal="center", vertical="center")
        cell3.border = self.thin_border
    
    # Summary and recommendations
    summary_row = dist_row + len(kanban_columns_order) + 3
    
    ws.merge_cells(f"A{summary_row}:F{summary_row}")
    summary_title = ws[f"A{summary_row}"]
    summary_title.value = "💡 REKOMENDASI & TINDAK LANJUT"
    summary_title.font = Font(bold=True, size=12)
    summary_title.alignment = Alignment(horizontal="center", vertical="center")
    
    summary_row += 1
    
    recommendations = []
    
    # Analyze for recommendations
    masalah_cards = kanban_data.get("⚠️ MASALAH JADWAL", [])
    if masalah_cards:
        high_priority_count = sum(1 for card in masalah_cards if card.get("priority") == "High")
        if high_priority_count > 0:
            recommendations.append(f"• Ada {high_priority_count} masalah prioritas TINGGI yang perlu segera ditangani")
        
        overload_count = sum(1 for card in masalah_cards if card.get("label") == "Overload")
        if overload_count > 0:
            recommendations.append(f"• {overload_count} slot Poleks melebihi batas maksimal ({self.max_e})")
        
        conflict_count = sum(1 for card in masalah_cards if card.get("label") == "Konflik")
        if conflict_count > 0:
            recommendations.append(f"• {conflict_count} konflik jadwal dokter ditemukan")
    
    penyesuaian_cards = kanban_data.get("🔧 PERLU PENYESUAIAN", [])
    if penyesuaian_cards:
        recommendations.append(f"• {len(penyesuaian_cards)} item perlu penyesuaian untuk optimalisasi")
    
    optimal_cards = kanban_data.get("✅ OPTIMAL", [])
    if optimal_cards:
        recommendations.append(f"• {len(optimal_cards)} pola jadwal sudah optimal - pertahankan")
    
    # Add default recommendations if none
    if not recommendations:
        recommendations = [
            "• Review distribusi dokter per poli",
            "• Periksa jam sibuk (10:00-12:00) untuk penambahan slot",
            "• Optimalkan penggunaan slot pagi vs sore",
            "• Verifikasi tidak ada konflik jadwal dokter"
        ]
    
    # Write recommendations
    for i, rec in enumerate(recommendations):
        ws.merge_cells(f"A{summary_row + i}:F{summary_row + i}")
        cell = ws[f"A{summary_row + i}"]
        cell.value = rec
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.font = Font(size=10)
    
    # Footer
    footer_row = summary_row + len(recommendations) + 2
    
    ws.merge_cells(f"A{footer_row}:F{footer_row}")
    footer = ws[f"A{footer_row}"]
    footer.value = "📌 Catatan: Laporan ini di-generate otomatis dari Kanban Board di aplikasi. Update secara berkala di tab Kanban."
    footer.font = Font(italic=True, size=9, color="666666")
    footer.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    
    return ws    
    def _generate_kanban_from_analysis(self, df_grid, slot_str):
        """Generate kanban data dari analisis jadwal jika tidak ada data kanban"""
        kanban_data = {
            "⚠️ MASALAH JADWAL": [],
            "🔧 PERLU PENYESUAIAN": [],
            "⏳ DALAM PROSES": [],
            "✅ OPTIMAL": []
        }
        
        if df_grid is None or df_grid.empty:
            # Default data jika tidak ada jadwal
            kanban_data["⚠️ MASALAH JADWAL"].append({
                "text": "Tidak ada data jadwal yang diproses",
                "label": "Kosong",
                "priority": "High"
            })
            return kanban_data
        
        try:
            # 1. Cek overload slots
            for hari in df_grid["HARI"].unique():
                hari_data = df_grid[df_grid["HARI"] == hari]
                
                for slot in slot_str[:15]:  # Check first 15 slots
                    if slot in hari_data.columns:
                        e_count = (hari_data[slot] == "E").sum()
                        
                        if e_count > self.max_e:
                            kanban_data["⚠️ MASALAH JADWAL"].append({
                                "text": f"{hari} {slot}: {e_count} Poleks (batas {self.max_e})",
                                "label": "Overload",
                                "priority": "High"
                            })
            
            # 2. Cek doctor conflicts
            for (dokter, hari), group in df_grid.groupby(["DOKTER", "HARI"]):
                if len(group) > 1:
                    conflict_slots = []
                    
                    for slot in slot_str[:10]:
                        if slot in group.columns:
                            active_polis = group[group[slot].isin(["R", "E"])]["POLI"].tolist()
                            if len(active_polis) > 1:
                                conflict_slots.append(slot)
                    
                    if conflict_slots:
                        kanban_data["⚠️ MASALAH JADWAL"].append({
                            "text": f"Dr. {dokter} - {hari}: konflik di {len(conflict_slots)} slot",
                            "label": "Konflik",
                            "priority": "High"
                        })
            
            # 3. Cek empty slots di peak hours
            peak_slots = [s for s in slot_str if "10:00" <= s <= "12:00"]
            
            for hari in df_grid["HARI"].unique():
                hari_data = df_grid[df_grid["HARI"] == hari]
                
                for poli in hari_data["POLI"].unique():
                    poli_data = hari_data[hari_data["POLI"] == poli]
                    
                    empty_in_peak = 0
                    for slot in peak_slots:
                        if slot in poli_data.columns:
                            slot_values = poli_data[slot].values
                            if not any(val in ["R", "E"] for val in slot_values):
                                empty_in_peak += 1
                    
                    if empty_in_peak >= 2:
                        kanban_data["🔧 PERLU PENYESUAIAN"].append({
                            "text": f"{poli} - {hari}: {empty_in_peak} slot kosong di jam sibuk",
                            "label": "Kosong",
                            "priority": "Medium"
                        })
            
            # 4. Cek distribusi
            for poli in df_grid["POLI"].unique():
                poli_data = df_grid[df_grid["POLI"] == poli]
                
                morning_slots = [s for s in slot_str if s < "12:00"]
                afternoon_slots = [s for s in slot_str if s >= "12:00"]
                
                morning_count = 0
                afternoon_count = 0
                
                for slot in morning_slots:
                    if slot in poli_data.columns:
                        morning_count += int((poli_data[slot] == "R").sum())
                        morning_count += int((poli_data[slot] == "E").sum())
                
                for slot in afternoon_slots:
                    if slot in poli_data.columns:
                        afternoon_count += int((poli_data[slot] == "R").sum())
                        afternoon_count += int((poli_data[slot] == "E").sum())
                
                total = morning_count + afternoon_count
                if total > 0:
                    morning_pct = (morning_count / total) * 100
                    
                    if morning_pct > 70:
                        kanban_data["🔧 PERLU PENYESUAIAN"].append({
                            "text": f"{poli}: {morning_pct:.0f}% pagi, terlalu dominan",
                            "label": "Distribusi",
                            "priority": "Medium"
                        })
            
            # 5. Temukan optimal schedules
            for hari in df_grid["HARI"].unique():
                hari_data = df_grid[df_grid["HARI"] == hari]
                
                for slot in ["10:00", "11:00", "13:00"]:
                    if slot in hari_data.columns:
                        doctor_count = (hari_data[slot].isin(["R", "E"])).sum()
                        
                        if 3 <= doctor_count <= 5:
                            kanban_data["✅ OPTIMAL"].append({
                                "text": f"{hari} {slot}: {int(doctor_count)} dokter (optimal)",
                                "label": "Optimal",
                                "priority": "Low"
                            })
            
            # 6. Default dalam proses
            kanban_data["⏳ DALAM PROSES"].append({
                "text": "Review jadwal Poli Anak",
                "label": "Review",
                "priority": "Low"
            })
            
        except Exception as e:
            print(f"⚠️ Error generating kanban data: {e}")
        
        return kanban_data
    
    def _format_kanban_card(self, card):
        """Format kanban card untuk ditampilkan di Excel"""
        text = card.get("text", "")
        label = card.get("label", "")
        priority = card.get("priority", "Medium")
        
        return f"{text}\n\nLabel: {label}\nPrioritas: {priority}"
    
    def _get_label_fill(self, label):
        """Get fill color untuk label"""
        label_fills = {
            "Overload": self.fill_overload,
            "Konflik": self.fill_konflik,
            "Kosong": self.fill_kosong,
            "Distribusi": self.fill_distribusi,
            "Beban": self.fill_beban,
            "Review": self.fill_review,
            "Optimal": self.fill_optimal
        }
        return label_fills.get(label)
    
    def _create_summary_sheet(self, wb, df_grid, slot_str):
        """Buat sheet Summary dengan statistik"""
        if "Summary" in wb.sheetnames:
            del wb["Summary"]
        
        ws = wb.create_sheet("Summary")
        
        # Title
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "SUMMARY JADWAL DOKTER"
        title_cell.font = Font(bold=True, size=16, color="366092")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Statistics
        stats = self._calculate_statistics(df_grid, slot_str)
        
        ws.append([])  # Empty row
        ws.append(["STATISTIK", "", "", ""])
        ws.append(["Total Baris Data", stats["total_rows"], "", ""])
        ws.append(["Total Dokter Unik", stats["total_doctors"], "", ""])
        ws.append(["Total Poli Unik", stats["total_poli"], "", ""])
        ws.append(["Total Slot Waktu", stats["total_slots"], "", ""])
        ws.append(["Slot Reguler (R)", stats["total_r"], f"{stats['total_r'] / stats['total_slots'] * 100:.1f}%" if stats['total_slots'] > 0 else "0%", ""])
        ws.append(["Slot Poleks (E)", stats["total_e"], f"{stats['total_e'] / stats['total_slots'] * 100:.1f}%" if stats['total_slots'] > 0 else "0%", ""])
        ws.append(["Slot Kosong", stats["total_empty"], f"{stats['total_empty'] / stats['total_slots'] * 100:.1f}%" if stats['total_slots'] > 0 else "0%", ""])
        ws.append(["Persentase Terisi", f"{stats['fill_percentage']:.1f}%", "", ""])
        
        # Configuration
        ws.append([])  # Empty row
        ws.append(["KONFIGURASI", "", "", ""])
        ws.append(["Jam Mulai", f"{self.config.start_hour:02d}:{self.config.start_minute:02d}", "", ""])
        ws.append(["Interval Slot", f"{self.config.interval_minutes} menit", "", ""])
        ws.append(["Maks Poleks/Slot", self.config.max_poleks_per_slot, "", ""])
        ws.append(["Auto Fix Errors", "Ya" if self.config.auto_fix_errors else "Tidak", "", ""])
        ws.append(["Hari Sabtu", "Aktif" if self.config.enable_sabtu else "Nonaktif", "", ""])
        
        # Poleks overload warning
        if df_grid is not None and not df_grid.empty:
            overload_count = self._count_poleks_overload(df_grid, slot_str)
            if overload_count > 0:
                ws.append([])
                ws.append(["PERINGATAN", f"{overload_count} slot Poleks melebihi batas!", "", ""])
        
        # Sheet list
        ws.append([])  # Empty row
        ws.append(["DAFTAR SHEET", "", "", ""])
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            if sheet_name != "Summary":  # Skip current sheet
                ws.append([f"{i-1}.", sheet_name, "", ""])
        
        # Kanban info
        ws.append([])  # Empty row
        ws.append(["KANBAN BOARD", "", "", ""])
        ws.append(["Sheet 'Kanban Board' berisi:", "", "", ""])
        ws.append(["• Analisis masalah jadwal", "", "", ""])
        ws.append(["• Prioritas perbaikan", "", "", ""])
        ws.append(["• Status penyelesaian", "", "", ""])
        
        # Timestamp
        ws.append([])  # Empty row
        ws.append(["Dibuat pada", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", ""])
        ws.append(["Aplikasi", "Sistem Jadwal Dokter v1.0", "", ""])
        
        # Style
        self._style_summary_sheet(ws)
        
        return ws
    
    def _count_poleks_overload(self, df_grid, slot_str):
        """Hitung berapa banyak slot yang melebihi batas poleks"""
        overload_count = 0
        
        if df_grid is None or df_grid.empty:
            return overload_count
        
        for hari in df_grid["HARI"].unique():
            hari_data = df_grid[df_grid["HARI"] == hari]
            
            for slot in slot_str:
                if slot in hari_data.columns:
                    e_count = (hari_data[slot] == "E").sum()
                    if e_count > self.max_e:
                        overload_count += 1
        
        return overload_count
    
    # ======================================================
    # STYLING METHODS UNTUK SHEET LAIN
    # ======================================================
    
    def _style_rekap_sheet(self, ws):
        """Style sheet rekap"""
        if ws.max_row <= 1:
            return
        
        # Header
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.fill_header
            cell.font = self.font_header
            cell.alignment = self.align_center
            cell.border = self.thin_border
        
        # Data rows
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                
                # Align numeric columns to right
                if col >= 3 and cell.value is not None:
                    try:
                        # Coba konversi ke float
                        if isinstance(cell.value, str) and '%' in cell.value:
                            cell.alignment = self.align_right
                        else:
                            float(str(cell.value).replace('%', ''))
                            cell.alignment = self.align_right
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                    except (ValueError, TypeError):
                        cell.alignment = self.align_left
                else:
                    cell.alignment = self.align_left
            
            # Alternate row colors
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = self.fill_gray
    
    def _style_conflict_sheet(self, ws):
        """Style conflict sheet"""
        self._style_rekap_sheet(ws)
        
        # Highlight conflict rows berdasarkan tingkat
        for row in range(2, ws.max_row + 1):
            tingkat_cell = ws.cell(row=row, column=5)  # Column E = TINGKAT
            tingkat = tingkat_cell.value if tingkat_cell.value else ""
            
            if "TINGGI" in str(tingkat).upper():
                fill_color = PatternFill("solid", fgColor="FFC7CE")  # Merah muda
            elif "SEDANG" in str(tingkat).upper():
                fill_color = PatternFill("solid", fgColor="FFE699")  # Kuning muda
            elif "RENDAH" in str(tingkat).upper():
                fill_color = PatternFill("solid", fgColor="C6EFCE")  # Hijau muda
            else:
                continue
            
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill_color
    
    def _style_conflict_map_sheet(self, ws, num_doctors):
        """Style conflict map sheet"""
        if ws.max_row <= 1:
            return
        
        # Header
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.fill_header
            cell.font = self.font_header
            cell.alignment = self.align_center
            cell.border = self.thick_border
        
        # Slot column (kolom pertama)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            cell.font = self.font_bold
            cell.alignment = self.align_center
            cell.border = self.thin_border
        
        # Doctor columns
        for row in range(2, ws.max_row + 1):
            for col in range(2, min(ws.max_column + 1, num_doctors + 2)):
                cell = ws.cell(row=row, column=col)
                cell.alignment = self.align_center
                cell.border = self.thin_border
                
                # Style untuk sel dengan konflik
                if cell.value in ["⚠️", "🚨"]:
                    cell.font = Font(size=12, bold=True)
    
    def _style_chart_sheet(self, ws):
        """Style chart sheet"""
        if ws.max_row <= 1:
            return
        
        # Header
        for col in range(1, 3):  # Hanya 2 kolom
            cell = ws.cell(row=1, column=col)
            cell.fill = self.fill_header
            cell.font = self.font_header
            cell.alignment = self.align_center
            cell.border = self.thin_border
        
        # Data
        for row in range(2, ws.max_row + 1):
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                cell.alignment = self.align_left if col == 1 else self.align_right
        
        # Alternate rows
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, 3):
                    ws.cell(row=row, column=col).fill = self.fill_gray
    
    def _style_summary_sheet(self, ws):
        """Style summary sheet"""
        if ws.max_row <= 1:
            return
        
        # Beri border pada semua sel
        for row in range(1, ws.max_row + 1):
            for col in range(1, 5):  # Hanya 4 kolom
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
        
        # Style title
        title_cell = ws["A1"]
        title_cell.fill = PatternFill("solid", fgColor="4F81BD")
        
        # Style section headers
        section_rows = [3, 14, 22, 30]  # Row numbers dengan section headers
        for row in section_rows:
            if row <= ws.max_row:
                cell = ws.cell(row=row, column=1)
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.font = Font(bold=True, size=11)
        
        # Style data rows
        for row in range(1, ws.max_row + 1):
            cell1 = ws.cell(row=row, column=1)
            cell2 = ws.cell(row=row, column=2)
            
            cell1.alignment = self.align_left
            cell2.alignment = self.align_left
            
            # Bold untuk statistik penting
            if row in [4, 5, 6, 10, 11, 12]:
                cell1.font = self.font_bold
                cell2.font = self.font_bold
        
        # Merge cells untuk title
        ws.merge_cells("A1:D1")
    
    def _apply_styling_to_all_sheets(self, wb):
        """Apply basic styling to all sheets"""
        for ws in wb.worksheets:
            # Set default font untuk semua sel
            for row in ws.iter_rows():
                for cell in row:
                    if cell.font is None or cell.font.name == 'Calibri':
                        cell.font = self.font_normal
    
    def _auto_adjust_column_widths(self, wb):
        """Auto adjust column widths untuk semua sheets"""
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            # Hitung panjang string
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Adjust width (min 10, max 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _reorder_sheets(self, wb):
        """Reorder sheets untuk UX yang lebih baik"""
        desired_order = [
            "Summary",
            "Jadwal", 
            "Rekap Layanan",
            "Rekap Poli",
            "Rekap Dokter",
            "Peak Hour Analysis",
            "Conflict Dokter",
            "Peta Konflik Dokter",
            "Grafik Poli",
            "Kanban Board"
        ]
        
        # Hanya reorder sheets yang ada
        existing_sheets = [s for s in desired_order if s in wb.sheetnames]
        
        # Pindahkan sheets ke posisi yang diinginkan
        for i, sheet_name in enumerate(existing_sheets):
            ws = wb[sheet_name]
            wb.move_sheet(ws, offset=-len(wb.sheetnames) + i)
    
    # ======================================================
    # HELPER METHODS
    # ======================================================
    
    def _combine_slots_to_ranges(self, slots, slot_str):
        """Gabungkan slot menjadi range waktu"""
        if not slots:
            return []
        
        try:
            # Pastikan slots ada dalam slot_str
            valid_slots = [s for s in slots if s in slot_str]
            if not valid_slots:
                return slots
            
            # Urutkan berdasarkan posisi di slot_str
            valid_slots.sort(key=lambda x: slot_str.index(x))
            
            ranges = []
            start = valid_slots[0]
            end = valid_slots[0]
            
            for i in range(1, len(valid_slots)):
                current_idx = slot_str.index(valid_slots[i])
                prev_idx = slot_str.index(end)
                
                # Cek jika berurutan
                if current_idx == prev_idx + 1:
                    end = valid_slots[i]
                else:
                    if start == end:
                        ranges.append(start)
                    else:
                        ranges.append(f"{start}-{end}")
                    
                    start = valid_slots[i]
                    end = valid_slots[i]
            
            # Tambahkan range terakhir
            if start == end:
                ranges.append(start)
            else:
                ranges.append(f"{start}-{end}")
            
            return ranges
            
        except Exception as e:
            print(f"⚠️ Error combining slots: {e}")
            return slots
    
    def _calculate_duration(self, time_range, slot_str):
        """Hitung durasi dalam jam dari range waktu"""
        try:
            if '-' in time_range:
                start_str, end_str = time_range.split('-')
                
                if start_str in slot_str and end_str in slot_str:
                    start_idx = slot_str.index(start_str)
                    end_idx = slot_str.index(end_str)
                    num_slots = end_idx - start_idx + 1
                    return num_slots * self.interval / 60
            
            # Fallback: single slot
            return self.interval / 60
            
        except:
            return 0
    
    def _find_doctor_conflicts(self, df_grid, slot_str):
        """Temukan konflik dokter"""
        conflicts = []
        
        if df_grid is None or df_grid.empty:
            return conflicts
        
        try:
            for (dokter, hari), group in df_grid.groupby(["DOKTER", "HARI"]):
                if len(group) > 1:  # Dokter muncul di >1 poli di hari yang sama
                    for slot in slot_str:
                        if slot in group.columns:
                            active_rows = group[group[slot].isin(["R", "E"])]
                            
                            if len(active_rows) > 1:
                                active_polis = active_rows["POLI"].tolist()
                                
                                # Cek untuk konflik R vs E
                                has_r = any(active_rows[slot] == "R")
                                has_e = any(active_rows[slot] == "E")
                                
                                if has_r and has_e:
                                    tingkat = "TINGGI"
                                    keterangan = f"Bentrok Reguler & Poleks"
                                else:
                                    tingkat = "SEDANG"
                                    keterangan = f"{len(active_polis)} poli bersamaan"
                                
                                conflicts.append({
                                    "dokter": dokter,
                                    "hari": hari,
                                    "slot": slot,
                                    "keterangan": keterangan,
                                    "tingkat": tingkat
                                })
        except Exception as e:
            print(f"⚠️ Error finding conflicts: {e}")
        
        return conflicts
    
    def _calculate_statistics(self, df_grid, slot_str):
        """Hitung statistik untuk summary"""
        stats = {
            "total_rows": 0,
            "total_doctors": 0,
            "total_poli": 0,
            "total_slots": 0,
            "total_r": 0,
            "total_e": 0,
            "total_empty": 0,
            "fill_percentage": 0
        }
        
        if df_grid is not None and not df_grid.empty:
            stats["total_rows"] = len(df_grid)
            stats["total_doctors"] = df_grid["DOKTER"].nunique()
            stats["total_poli"] = df_grid["POLI"].nunique()
            
            total_cells = len(df_grid) * len(slot_str)
            stats["total_slots"] = total_cells
            
            # Hitung R dan E
            for slot in slot_str:
                if slot in df_grid.columns:
                    stats["total_r"] += (df_grid[slot] == "R").sum()
                    stats["total_e"] += (df_grid[slot] == "E").sum()
            
            stats["total_empty"] = total_cells - stats["total_r"] - stats["total_e"]
            
            if total_cells > 0:
                stats["fill_percentage"] = ((stats["total_r"] + stats["total_e"]) / total_cells) * 100
        
        return stats
    
    # ======================================================
    # TEMPLATE GENERATOR
    # ======================================================
    
    def generate_template(self, slot_str=None):
        """
        Generate template Excel file untuk input
        
        Returns:
            BytesIO buffer berisi template Excel
        """
        print("📄 Generating template Excel...")
        
        try:
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            
            # Generate slot strings jika tidak provided
            if slot_str is None:
                from datetime import time
                slot_str = []
                current_time = self.config.start_hour * 60 + self.config.start_minute
                end_time = 14 * 60 + 30  # 14:30
                
                while current_time < end_time:
                    hours = current_time // 60
                    minutes = current_time % 60
                    time_str = f"{hours:02d}:{minutes:02d}"
                    slot_str.append(time_str)
                    current_time += self.config.interval_minutes
            
            # Create Reguler sheet
            ws_reg = wb.create_sheet("Reguler")
            self._create_template_sheet(ws_reg, "Reguler", slot_str)
            
            # Create Poleks sheet
            ws_pol = wb.create_sheet("Poleks")
            self._create_template_sheet(ws_pol, "Poleks", slot_str)
            
            # Create Instructions sheet
            ws_inst = wb.create_sheet("Instruksi")
            self._create_instructions_sheet(ws_inst)
            
            # Apply styling
            self._apply_styling_to_all_sheets(wb)
            self._auto_adjust_column_widths(wb)
            
            # Reorder sheets
            self._reorder_sheets(wb)
            
            # Save to buffer
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            
            file_size = buf.getbuffer().nbytes
            print(f"✅ Template created: {file_size:,} bytes")
            return buf
            
        except Exception as e:
            print(f"❌ Error generating template: {e}")
            print(traceback.format_exc())
            
            # Fallback template sederhana
            wb = Workbook()
            ws = wb.active
            ws.title = "Template"
            ws.append(["Nama Dokter", "Poli Asal", "Senin", "Selasa", "Rabu", "Kamis", "Jum'at"])
            ws.append(["dr. Contoh", "Poli Anak", "08.00-10.00", "09.00-11.00", "", "10.00-12.00", "08.00-10.00"])
            
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf
    
    def _create_template_sheet(self, ws, jenis, slot_str):
        """Buat sheet template untuk Reguler atau Poleks"""
        # Header
        headers = ["Nama Dokter", "Poli Asal", "Jenis Poli", "Senin", "Selasa", "Rabu", "Kamis", "Jum'at"]
        
        if self.config.enable_sabtu:
            headers.append("Sabtu")
        
        ws.append(headers)
        
        # Example data
        examples = [
            ["dr. Contoh Satu", "Poli Anak", jenis, "08.00-10.00", "09.00-11.00", "", "10.00-12.00", "08.00-10.00"],
            ["dr. Contoh Dua", "Poli Dalam", jenis, "10.00-12.00", "", "08.00-10.00", "09.00-11.00", ""],
        ]
        
        if self.config.enable_sabtu:
            for ex in examples:
                ex.append("")  # Add empty Sabtu column
        
        for example in examples:
            ws.append(example)
        
        # Notes
        notes_row = len(examples) + 2
        ws.cell(row=notes_row, column=1, value="CATATAN:")
        ws.cell(row=notes_row + 1, column=1, value="1. Format waktu: '07.30-10.00' atau '07:30-10:00'")
        ws.cell(row=notes_row + 2, column=1, value="2. Kosongkan jika tidak ada jadwal")
        ws.cell(row=notes_row + 3, column=1, value="3. Jenis Poli akan otomatis terisi")
        
        # Style header
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.fill_header
            cell.font = self.font_header
            cell.alignment = self.align_center
        
        # Style example rows
        for row in range(2, 4):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
        
        ws.freeze_panes = "D2"
    
    def _create_instructions_sheet(self, ws):
        """Buat sheet instruksi"""
        instructions = [
            "PETUNJUK PENGGUNAAN",
            "",
            "1. FILE INPUT:",
            "   - File Excel harus memiliki 2 sheet: 'Reguler' dan 'Poleks'",
            "   - Format kolom harus sama seperti di template",
            "",
            "2. FORMAT WAKTU:",
            "   - Gunakan format: '07.30-10.00' atau '07:30-10:00'",
            "   - Bisa menggunakan titik atau titik dua",
            "   - Kosongkan sel jika tidak ada jadwal",
            "",
            "3. KOLOM WAJIB:",
            "   - Nama Dokter: Nama lengkap dokter",
            "   - Poli Asal: Nama poli",
            "   - Jenis Poli: Akan otomatis terisi",
            "   - Kolom hari: Senin sampai Jum'at",
            "",
            "4. PROSES:",
            "   - Upload file di tab 'Upload & Proses'",
            "   - Klik 'Proses Jadwal' untuk konversi",
            "   - Download hasil di file Excel lengkap",
            "",
            "5. OUTPUT:",
            "   - File hasil berisi 10+ sheet dengan analisis lengkap",
            "   - Termasuk deteksi konflik, statistik, dan Kanban Board",
            "",
            "6. WARNA PADA SHEET JADWAL:",
            "   - HIJAU: Jadwal Reguler (R)",
            "   - BIRU: Jadwal Poleks dalam batas",
            "   - MERAH: Jadwal Poleks melebihi batas",
            "",
            "7. KANBAN BOARD:",
            "   - Sheet khusus untuk manajemen masalah jadwal",
            "   - Prioritas: Merah (Tinggi), Kuning (Sedang), Hijau (Rendah)",
            "   - Kolom: Masalah, Perlu Penyesuaian, Dalam Proses, Optimal",
            "",
            "© 2024 Sistem Jadwal Dokter"
        ]
        
        for i, line in enumerate(instructions, start=1):
            ws.cell(row=i, column=1, value=line)
            if line.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.")):
                ws.cell(row=i, column=1).font = Font(bold=True)
        
        # Title style
        title_cell = ws["A1"]
        title_cell.font = Font(bold=True, size=14, color="366092")
        
        ws.column_dimensions['A'].width = 80
