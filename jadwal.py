import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import time, timedelta, datetime
import io
import re
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
from typing import Dict, List, Tuple, Optional, Any
import traceback
import warnings

# Suppress warnings
warnings.filterwarnings('ignore')

# ============================================================================
# KELAS UTAMA - Aplikasi Pengisi Jadwal Poli
# ============================================================================

class PoliSchedulerApp:
    """Kelas utama aplikasi Pengisi Jadwal Poli"""
    
    def __init__(self):
        """Inisialisasi aplikasi"""
        self.setup_page_config()
        self.initialize_session_state()
        self.initialize_constants()
        
    def setup_page_config(self):
        """Setup konfigurasi halaman Streamlit"""
        st.set_page_config(
            page_title="🏥 Pengisi Jadwal Poli Excel - OOP Version",
            page_icon="🏥",
            layout="wide",
            initial_sidebar_state="expanded"
        )
        
    def initialize_session_state(self):
        """Inisialisasi session state"""
        session_defaults = {
            'processed_data': None,
            'error_logs': [],
            'processing_stats': {},
            'visualization_data': None,
            'uploaded_file_name': None,
            'last_processed_time': None,
            'config_changed': False
        }
        
        for key, default in session_defaults.items():
            if key not in st.session_state:
                st.session_state[key] = default
                
    def initialize_constants(self):
        """Inisialisasi konstan"""
        # Warna untuk sel
        self.FILL_R = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        self.FILL_E = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        self.FILL_OVERLIMIT = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Hari
        self.HARI_ORDER = {"Senin": 1, "Selasa": 2, "Rabu": 3, "Kamis": 4, "Jum'at": 5}
        self.HARI_INDONESIA = ["Senin", "Selasa", "Rabu", "Kamis", "Jum'at"]
        
        # Konfigurasi default
        self.config = {
            'start_hour': 7,
            'start_minute': 30,
            'interval_minutes': 30,
            'max_poleks_per_slot': 7,
            'auto_fix_errors': True,
            'color_theme': 'Default'
        }
        
        # Generate time slots based on config
        self._update_time_slots()
    
    def _update_time_slots(self):
        """Update time slots berdasarkan konfigurasi"""
        slots = []
        current_time = time(
            self.config['start_hour'],
            self.config['start_minute']
        )
        
        interval = self.config['interval_minutes']
        end_time = time(14, 30)  # Batas akhir
        
        while current_time <= end_time:
            slots.append(current_time)
            
            # Tambah interval
            current_datetime = datetime.combine(datetime.today(), current_time)
            next_datetime = current_datetime + timedelta(minutes=interval)
            current_time = next_datetime.time()
        
        self.TIME_SLOTS = slots
        self.TIME_SLOTS_STR = [t.strftime("%H:%M") for t in slots]


# ============================================================================
# KELAS UTILITAS
# ============================================================================

class ErrorAnalyzer:
    """Kelas untuk menganalisis dan menangani error"""
    
    @staticmethod
    def analyze_excel_structure(df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Analisis struktur dataframe Excel"""
        errors = []
        warnings_list = []
        
        # Cek kolom yang diperlukan
        required_columns = ['Nama Dokter', 'Poli Asal', 'Jenis Poli']
        hari_columns = ['Senin', 'Selasa', 'Rabu', 'Kamis', "Jum'at"]
        
        missing_required = [col for col in required_columns if col not in df.columns]
        if missing_required:
            errors.append(f"Kolom yang diperlukan tidak ditemukan: {missing_required}")
        
        missing_hari = [col for col in hari_columns if col not in df.columns]
        if missing_hari:
            warnings_list.append(f"Kolom hari yang tidak ditemukan: {missing_hari}")
        
        # Cek data kosong
        empty_rows = df[required_columns].isnull().all(axis=1).sum()
        if empty_rows > 0:
            warnings_list.append(f"{empty_rows} baris kosong ditemukan")
        
        # Cek format waktu
        time_format_errors = 0
        invalid_times = []
        for hari in hari_columns:
            if hari in df.columns:
                for idx, time_str in enumerate(df[hari].dropna(), start=2):
                    if not ErrorAnalyzer._validate_time_format(str(time_str)):
                        time_format_errors += 1
                        invalid_times.append(f"Baris {idx}: '{time_str}'")
        
        if time_format_errors > 0:
            errors.append(f"{time_format_errors} format waktu tidak valid ditemukan")
            if time_format_errors <= 10:  # Batasi output jika terlalu banyak
                errors.extend(invalid_times)
        
        # Cek duplikat dokter pada hari yang sama
        if all(col in df.columns for col in required_columns[:2] + hari_columns[:1]):
            duplicates = df.groupby(['Nama Dokter', 'Poli Asal', 'Senin']).filter(lambda x: len(x) > 1)
            if len(duplicates) > 0:
                warnings_list.append(f"{len(duplicates)} kemungkinan duplikat jadwal ditemukan")
        
        return {
            'is_valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings_list,
            'total_rows': len(df),
            'missing_columns': missing_required + missing_hari,
            'time_format_errors': time_format_errors
        }
    
    @staticmethod
    def _validate_time_format(time_str: str) -> bool:
        """Validasi format waktu"""
        if pd.isna(time_str) or str(time_str).strip() == "":
            return True
        
        clean_str = str(time_str).strip()
        
        # Pattern untuk format waktu yang diterima
        patterns = [
            r'^\d{1,2}\.\d{2}\s*-\s*\d{1,2}\.\d{2}$',  # 08.00 - 10.30
            r'^\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}$',    # 08:00 - 10:30
            r'^\d{1,2}\.\d{2}-\d{1,2}\.\d{2}$',        # 08.00-10.30
            r'^\d{1,2}:\d{2}-\d{1,2}:\d{2}$',          # 08:00-10:30
            r'^\d{1,2}\.\d{2}\s*-\s*\d{1,2}:\d{2}$',   # Campuran
            r'^\d{1,2}:\d{2}\s*-\s*\d{1,2}\.\d{2}$',   # Campuran
        ]
        
        # Cek pattern
        if any(re.match(pattern, clean_str) for pattern in patterns):
            # Validasi jam dan menit
            try:
                # Ekstrak waktu
                times = re.split(r'\s*-\s*', clean_str.replace(':', '.'))
                if len(times) == 2:
                    start_hour, start_minute = map(float, times[0].split('.'))
                    end_hour, end_minute = map(float, times[1].split('.'))
                    
                    # Validasi rentang
                    if (0 <= start_hour < 24 and 0 <= end_hour < 24 and
                        0 <= start_minute < 60 and 0 <= end_minute < 60):
                        return True
            except:
                return False
        
        return False
    
    @staticmethod
    def create_error_report(error_data: Dict[str, Any]) -> str:
        """Buat laporan error"""
        report = "## 📋 Laporan Analisis Error\n\n"
        
        if error_data['errors']:
            report += "### ❌ Errors:\n"
            for error in error_data['errors'][:15]:  # Batasi output
                report += f"- {error}\n"
            if len(error_data['errors']) > 15:
                report += f"- ... dan {len(error_data['errors']) - 15} error lainnya\n"
            report += "\n"
        
        if error_data['warnings']:
            report += "### ⚠️ Warnings:\n"
            for warning in error_data['warnings']:
                report += f"- {warning}\n"
            report += "\n"
        
        report += f"### 📊 Statistik:\n"
        report += f"- Total Baris: {error_data['total_rows']}\n"
        report += f"- Format Waktu Error: {error_data['time_format_errors']}\n"
        report += f"- Kolom Hilang: {len(error_data['missing_columns'])}\n"
        report += f"- Status Valid: {'✅ Ya' if error_data['is_valid'] else '❌ Tidak'}\n"
        
        return report


class DataProcessor:
    """Kelas untuk memproses data jadwal"""
    
    def __init__(self, app: 'PoliSchedulerApp'):
        self.app = app
        self.time_parser = TimeParser()
        self.data_cleaner = DataCleaner()
        
    def process_schedule(self, df: pd.DataFrame, jenis_poli: str) -> pd.DataFrame:
        """Proses dataframe jadwal"""
        # Bersihkan data terlebih dahulu
        df_cleaned = self.data_cleaner.clean_data(df, jenis_poli)
        
        if df_cleaned.empty:
            return pd.DataFrame()
        
        results = []
        
        for (dokter, poli_asal), group in df_cleaned.groupby(['Nama Dokter', 'Poli Asal']):
            hari_schedules = self._extract_schedules_per_day(group)
            
            # Generate baris per hari
            for hari in self.app.HARI_INDONESIA:
                if hari not in hari_schedules or not hari_schedules[hari]:
                    continue
                    
                merged_ranges = self._merge_time_ranges(hari_schedules[hari])
                if merged_ranges:  # Hanya tambahkan jika ada jadwal
                    row = self._create_schedule_row(poli_asal, jenis_poli, hari, dokter, merged_ranges)
                    results.append(row)
        
        return pd.DataFrame(results) if results else pd.DataFrame()
    
    def _extract_schedules_per_day(self, group: pd.DataFrame) -> Dict[str, List[Tuple[time, time]]]:
        """Ekstrak jadwal per hari"""
        hari_schedules = {hari: [] for hari in self.app.HARI_INDONESIA}
        
        for hari in self.app.HARI_INDONESIA:
            if hari not in group.columns:
                continue
                
            time_ranges = []
            for time_str in group[hari]:
                start_time, end_time = self.time_parser.parse(time_str)
                if start_time and end_time:
                    # Potong maksimal sampai 14:30
                    if end_time > time(14, 30):
                        end_time = time(14, 30)
                    if start_time < end_time:  # Pastikan start < end
                        time_ranges.append((start_time, end_time))
            
            if time_ranges:
                hari_schedules[hari] = time_ranges
        
        return hari_schedules
    
    def _merge_time_ranges(self, time_ranges: List[Tuple[time, time]]) -> List[List[time]]:
        """Gabungkan rentang waktu yang overlapping"""
        if not time_ranges:
            return []
        
        # Sort by start time
        sorted_ranges = sorted(time_ranges, key=lambda x: x[0])
        merged = []
        
        for start, end in sorted_ranges:
            if not merged:
                merged.append([start, end])
            else:
                last_start, last_end = merged[-1]
                if start <= last_end:  # Overlap
                    merged[-1][1] = max(last_end, end)
                else:
                    merged.append([start, end])
        
        return merged
    
    def _create_schedule_row(self, poli_asal: str, jenis_poli: str, 
                            hari: str, dokter: str, 
                            merged_ranges: List[List[time]]) -> Dict[str, Any]:
        """Buat baris jadwal"""
        row = {
            'POLI ASAL': poli_asal,
            'JENIS POLI': jenis_poli,
            'HARI': hari,
            'DOKTER': dokter
        }
        
        # Isi slot waktu
        for i, slot_time in enumerate(self.app.TIME_SLOTS):
            slot_start = slot_time
            slot_end = (datetime.combine(datetime.today(), slot_start) + 
                       timedelta(minutes=self.app.config['interval_minutes'])).time()
            
            # Cek overlap dengan semua rentang waktu
            has_overlap = False
            for start, end in merged_ranges:
                if not (slot_end <= start or slot_start >= end):
                    has_overlap = True
                    break
            
            if has_overlap:
                row[self.app.TIME_SLOTS_STR[i]] = 'R' if jenis_poli == 'Reguler' else 'E'
            else:
                row[self.app.TIME_SLOTS_STR[i]] = ''
        
        return row


class TimeParser:
    """Kelas untuk parsing waktu"""
    
    @staticmethod
    def parse(time_str: Any) -> Tuple[Optional[time], Optional[time]]:
        """Parse rentang waktu dari string"""
        if pd.isna(time_str) or str(time_str).strip() == "":
            return None, None
        
        clean_str = str(time_str).strip()
        
        # Normalisasi format - ganti titik dengan titik dua
        clean_str = clean_str.replace(' ', '').replace('.', ':')
        
        # Cari pola waktu
        pattern = r'(\d{1,2}:\d{2})-(\d{1,2}:\d{2})'
        match = re.search(pattern, clean_str)
        
        if not match:
            # Coba pattern alternatif
            pattern2 = r'(\d{1,2})(\d{2})-(\d{1,2})(\d{2})'
            match2 = re.search(pattern2, clean_str)
            if match2:
                groups = match2.groups()
                start_str = f"{groups[0]}:{groups[1]}"
                end_str = f"{groups[2]}:{groups[3]}"
            else:
                return None, None
        else:
            start_str, end_str = match.groups()
        
        try:
            start_hour, start_minute = map(int, start_str.split(':'))
            end_hour, end_minute = map(int, end_str.split(':'))
            
            # Validasi waktu
            if not (0 <= start_hour < 24 and 0 <= end_hour < 24 and
                    0 <= start_minute < 60 and 0 <= end_minute < 60):
                return None, None
            
            start_time = time(start_hour, start_minute)
            end_time = time(end_hour, end_minute)
            
            # Jika end time lebih kecil dari start time (melewati tengah malam)
            if end_time < start_time:
                end_time = time(end_hour + 24, end_minute)
            
            return start_time, end_time
        except (ValueError, TypeError):
            return None, None


class DataCleaner:
    """Kelas untuk membersihkan data input"""
    
    @staticmethod
    def clean_data(df: pd.DataFrame, jenis_poli: str) -> pd.DataFrame:
        """Bersihkan dan perbaiki data input"""
        df_clean = df.copy()
        
        # Pastikan kolom yang diperlukan ada
        required_cols = ['Nama Dokter', 'Poli Asal', 'Jenis Poli']
        for col in required_cols:
            if col not in df_clean.columns:
                df_clean[col] = ''
        
        # Isi jenis poli jika kosong
        df_clean['Jenis Poli'] = df_clean['Jenis Poli'].fillna(jenis_poli)
        
        # Perbaiki format waktu
        hari_cols = ["Senin", "Selasa", "Rabu", "Kamis", "Jum'at"]
        for col in hari_cols:
            if col in df_clean.columns:
                df_clean[col] = df_clean[col].apply(DataCleaner._fix_time_format)
        
        # Hapus baris yang tidak memiliki jadwal sama sekali
        if hari_cols[0] in df_clean.columns:
            has_schedule = df_clean[hari_cols].notna().any(axis=1)
            df_clean = df_clean[has_schedule]
        
        # Hapus duplikat berdasarkan kolom utama
        df_clean = df_clean.drop_duplicates(subset=['Nama Dokter', 'Poli Asal'] + hari_cols)
        
        return df_clean.reset_index(drop=True)
    
    @staticmethod
    def _fix_time_format(time_str: Any) -> str:
        """Perbaiki format waktu yang tidak standar"""
        if pd.isna(time_str) or str(time_str).strip() == "":
            return ""
        
        str_time = str(time_str).strip()
        
        # Hapus karakter yang tidak perlu, tapi pertahankan spasi di sekitar '-'
        str_time = re.sub(r'[^\d.\-:\s]', '', str_time)
        
        # Normalisasi . menjadi :
        str_time = str_time.replace('.', ':')
        
        # Pastikan format HH:MM-HH:MM
        pattern = r'(\d{1,2}):?(\d{2})\s*-\s*(\d{1,2}):?(\d{2})'
        match = re.search(pattern, str_time)
        
        if match:
            start_hour, start_min, end_hour, end_min = match.groups()
            # Format ulang ke HH:MM-HH:MM
            return f"{start_hour}:{start_min}-{end_hour}:{end_min}"
        
        return str_time


class ExcelStyler:
    """Kelas untuk styling Excel"""
    
    def __init__(self, app: 'PoliSchedulerApp'):
        self.app = app
        
    def apply_styles(self, ws, max_row: int):
        """Terapkan styling ke sheet Jadwal"""
        # Hitung jumlah E per hari per slot
        e_counts = self._count_poleks_per_slot(ws, max_row)
        
        # Terapkan warna dan highlight
        self._apply_colors_and_highlights(ws, max_row, e_counts)
        
        # Auto-width columns
        self._auto_adjust_column_widths(ws)
    
    def _count_poleks_per_slot(self, ws, max_row: int) -> Dict[str, Dict[str, int]]:
        """Hitung jumlah Poleks per hari per slot"""
        e_counts = {hari: {slot: 0 for slot in self.app.TIME_SLOTS_STR} 
                   for hari in self.app.HARI_INDONESIA}
        
        for row in range(2, max_row + 1):
            hari = ws.cell(row=row, column=3).value
            if hari not in e_counts:
                continue
                
            for col_idx, slot in enumerate(self.app.TIME_SLOTS_STR, start=5):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value == 'E':
                    e_counts[hari][slot] += 1
        
        return e_counts
    
    def _apply_colors_and_highlights(self, ws, max_row: int, e_counts: Dict[str, Dict[str, int]]):
        """Terapkan warna dan highlight jika melebihi batas"""
        for row in range(2, max_row + 1):
            hari = ws.cell(row=row, column=3).value
            if hari not in e_counts:
                continue
                
            for col_idx, slot in enumerate(self.app.TIME_SLOTS_STR, start=5):
                cell = ws.cell(row=row, column=col_idx)
                
                if cell.value == 'R':
                    cell.fill = self.app.FILL_R
                elif cell.value == 'E':
                    cell.fill = self.app.FILL_E
                    
                    # Cek apakah melebihi batas
                    if e_counts[hari][slot] > self.app.config['max_poleks_per_slot']:
                        # Hitung berapa E yang sudah ada sebelum baris ini
                        e_count_before = 0
                        for r in range(2, row):
                            if (ws.cell(row=r, column=3).value == hari and 
                                ws.cell(row=r, column=col_idx).value == 'E'):
                                e_count_before += 1
                        
                        # Jika baris ini termasuk yang melebihi batas
                        if e_count_before >= self.app.config['max_poleks_per_slot']:
                            cell.fill = self.app.FILL_OVERLIMIT
    
    def _auto_adjust_column_widths(self, ws):
        """Auto-adjust column widths untuk keterbacaan"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width


class Visualizer:
    """Kelas untuk visualisasi data"""
    
    def __init__(self, app: 'PoliSchedulerApp'):
        self.app = app
        
    def create_heatmap(self, df: pd.DataFrame) -> go.Figure:
        """Buat heatmap jadwal"""
        if df.empty:
            return go.Figure()
        
        # Filter hanya kolom waktu
        time_cols = [col for col in df.columns if col in self.app.TIME_SLOTS_STR]
        
        # Siapkan data untuk heatmap
        pivot_data = []
        
        for _, row in df.iterrows():
            for slot in time_cols:
                value = row.get(slot, '')
                if value == 'R':
                    num_value = 1
                elif value == 'E':
                    num_value = 2
                else:
                    num_value = 0
                
                pivot_data.append({
                    'Hari': row['HARI'],
                    'Waktu': slot,
                    'Poli': row['POLI ASAL'],
                    'Dokter': row['DOKTER'][:20],  # Potong jika terlalu panjang
                    'Value': num_value,
                    'Label': value
                })
        
        heatmap_df = pd.DataFrame(pivot_data)
        
        if heatmap_df.empty:
            return go.Figure()
        
        # Buat pivot table untuk heatmap
        pivot_table = heatmap_df.pivot_table(
            index='Hari',
            columns='Waktu',
            values='Value',
            aggfunc='max',
            fill_value=0
        )
        
        # Urutkan hari
        hari_order = self.app.HARI_INDONESIA
        pivot_table = pivot_table.reindex([h for h in hari_order if h in pivot_table.index])
        
        # Buat heatmap
        fig = px.imshow(
            pivot_table,
            labels=dict(x="Waktu", y="Hari", color="Status"),
            x=pivot_table.columns,
            y=pivot_table.index,
            aspect="auto",
            color_continuous_scale=["white", "green", "blue", "red"],
            text_auto=False,
            title="Heatmap Jadwal Poli"
        )
        
        # Update layout
        fig.update_layout(
            height=max(400, len(pivot_table.index) * 40),
            xaxis_title="Waktu",
            yaxis_title="Hari",
            coloraxis_showscale=False
        )
        
        # Tambahkan annotation untuk nilai
        for i, hari in enumerate(pivot_table.index):
            for j, waktu in enumerate(pivot_table.columns):
                value = pivot_table.iloc[i, j]
                if value > 0:
                    fig.add_annotation(
                        x=j,
                        y=i,
                        text="R" if value == 1 else "E",
                        showarrow=False,
                        font=dict(color="white" if value > 0 else "black")
                    )
        
        return fig
    
    def create_schedule_gantt(self, df: pd.DataFrame) -> go.Figure:
        """Buat Gantt chart untuk jadwal"""
        if df.empty:
            return go.Figure()
        
        # Siapkan data untuk Gantt chart
        gantt_data = []
        
        for _, row in df.iterrows():
            for i, slot in enumerate(self.app.TIME_SLOTS_STR):
                if slot in df.columns and row[slot] in ['R', 'E']:
                    # Konversi waktu ke datetime untuk plotting
                    start_time = datetime.strptime(slot, "%H:%M")
                    end_time = start_time + timedelta(minutes=self.app.config['interval_minutes'])
                    
                    gantt_data.append({
                        'Task': f"{row['POLI ASAL']}\n{row['DOKTER'][:15]}...",
                        'Start': start_time,
                        'Finish': end_time,
                        'Resource': 'Reguler' if row[slot] == 'R' else 'Poleks',
                        'Hari': row['HARI'],
                        'Poli': row['POLI ASAL'],
                        'Jenis': row[slot]
                    })
        
        if not gantt_data:
            return go.Figure()
        
        gantt_df = pd.DataFrame(gantt_data)
        
        # Buat Gantt chart
        fig = px.timeline(
            gantt_df,
            x_start="Start",
            x_end="Finish",
            y="Task",
            color="Resource",
            facet_row="Hari",
            title="Gantt Chart Jadwal Poli",
            color_discrete_map={'Reguler': 'green', 'Poleks': 'blue'}
        )
        
        fig.update_layout(
            height=min(800, len(gantt_df['Task'].unique()) * 30 + 200),
            showlegend=True,
            xaxis_title="Waktu",
            yaxis_title=""
        )
        
        # Format x-axis
        fig.update_xaxes(
            tickformat="%H:%M",
            dtick=3600000  # 1 hour in milliseconds
        )
        
        return fig
    
    def create_statistics_charts(self, df: pd.DataFrame) -> Tuple[go.Figure, go.Figure]:
        """Buat chart statistik"""
        if df.empty:
            return go.Figure(), go.Figure()
        
        time_cols = [col for col in df.columns if col in self.app.TIME_SLOTS_STR]
        
        # Hitung statistik
        total_r = (df[time_cols] == 'R').sum().sum()
        total_e = (df[time_cols] == 'E').sum().sum()
        total_empty = (df[time_cols] == '').sum().sum()
        total_slots = len(df) * len(time_cols)
        
        # Pie chart untuk distribusi slot
        fig_pie = go.Figure(data=[go.Pie(
            labels=['Reguler', 'Poleks', 'Kosong'],
            values=[total_r, total_e, total_empty],
            hole=.3,
            marker_colors=['green', 'blue', 'lightgray'],
            textinfo='percent+value',
            hoverinfo='label+percent+value'
        )])
        
        fig_pie.update_layout(
            title=f"Distribusi Slot Waktu (Total: {total_slots})",
            height=350,
            showlegend=True
        )
        
        # Bar chart untuk slot per hari
        hari_stats = []
        for hari in self.app.HARI_INDONESIA:
            hari_df = df[df['HARI'] == hari]
            if not hari_df.empty:
                hari_r = (hari_df[time_cols] == 'R').sum().sum()
                hari_e = (hari_df[time_cols] == 'E').sum().sum()
                hari_stats.append({
                    'Hari': hari,
                    'Reguler': hari_r,
                    'Poleks': hari_e,
                    'Total': hari_r + hari_e
                })
        
        if hari_stats:
            stats_df = pd.DataFrame(hari_stats)
            
            fig_bar = go.Figure(data=[
                go.Bar(name='Reguler', x=stats_df['Hari'], y=stats_df['Reguler'], 
                       marker_color='green', text=stats_df['Reguler'], textposition='auto'),
                go.Bar(name='Poleks', x=stats_df['Hari'], y=stats_df['Poleks'], 
                       marker_color='blue', text=stats_df['Poleks'], textposition='auto')
            ])
            
            fig_bar.update_layout(
                title="Slot per Hari",
                barmode='stack',
                height=350,
                xaxis_title="Hari",
                yaxis_title="Jumlah Slot",
                showlegend=True
            )
        else:
            fig_bar = go.Figure()
        
        return fig_pie, fig_bar


# ============================================================================
# KELAS TAMPILAN (VIEW)
# ============================================================================

class TemplateManager:
    """Kelas untuk mengelola template Excel"""
    
    def __init__(self, app: 'PoliSchedulerApp'):
        self.app = app
    
    def create_template(self) -> io.BytesIO:
        """Buat template Excel"""
        wb = openpyxl.Workbook()
        
        # Sheet Poli Asal
        ws1 = wb.active
        ws1.title = "Poli Asal"
        ws1.append(["No", "Nama Poli", "kode sheet"])
        ws1.append(["", "", "(untuk referensi)"])
        
        poli_list = [
            ["1", "Poli Anak", "ANAK"],
            ["2", "Poli Bedah", "BEDAH"],
            ["3", "Poli Dalam", "DALAM"],
            ["4", "Poli Obgyn", "OBGYN"],
            ["5", "Poli Jantung", "JANTUNG"],
            ["6", "Poli Ortho", "ORTHO"],
            ["7", "Poli Paru", "PARU"],
            ["8", "Poli Saraf", "SARAF"],
            ["9", "Poli THT", "THT"],
            ["10", "Poli Urologi", "URO"],
            ["11", "Poli Jiwa", "JIWA"],
            ["12", "Poli Kukel", "KUKEL"],
            ["13", "Poli Bedah Saraf", "BSARAF"],
            ["14", "Poli Gigi", "GIGI"],
            ["15", "Poli Mata", "MATA"],
            ["16", "Poli Rehab", "REHAB"]
        ]
        for poli in poli_list:
            ws1.append(poli)
        
        # Format header
        for col in ws1.iter_cols(min_row=1, max_row=2, max_col=3):
            for cell in col:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = openpyxl.styles.Font(bold=True)
        
        # Sheet Reguler
        ws2 = wb.create_sheet("Reguler")
        headers = ["Nama Dokter", "Poli Asal", "Jenis Poli", "Senin", "Selasa", 
                  "Rabu", "Kamis", "Jum'at"]
        ws2.append(headers)
        
        # Contoh data
        contoh_data = [
            ["dr. Contoh Dokter, Sp.A", "Poli Anak", "Reguler", 
             "08.00 - 10.30", "", "08.00 - 10.30", "", ""],
            ["dr. Contoh Lain, Sp.PD", "Poli Dalam", "Reguler", 
             "", "09.00 - 12.00", "", "09.00 - 12.00", ""]
        ]
        for data in contoh_data:
            ws2.append(data)
        
        # Format header Reguler
        for col in ws2.iter_cols(min_row=1, max_row=1, max_col=len(headers)):
            for cell in col:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = openpyxl.styles.Font(bold=True)
        
        # Sheet Poleks
        ws3 = wb.create_sheet("Poleks")
        ws3.append(headers)
        
        # Contoh data Poleks
        contoh_poleks = [
            ["dr. Contoh Dokter, Sp.A", "Poli Anak", "Poleks", 
             "07.30 - 08.25", "", "07.30 - 08.25", "", ""],
            ["dr. Contoh Lain, Sp.PD", "Poli Dalam", "Poleks", 
             "", "08.00 - 09.00", "", "08.00 - 09.00", "10.00 - 11.00"]
        ]
        for data in contoh_poleks:
            ws3.append(data)
        
        # Format header Poleks
        for col in ws3.iter_cols(min_row=1, max_row=1, max_col=len(headers)):
            for cell in col:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                cell.font = openpyxl.styles.Font(bold=True)
        
        # Sheet Jadwal
        ws4 = wb.create_sheet("Jadwal")
        jadwal_headers = ["POLI ASAL", "JENIS POLI", "HARI", "DOKTER"] + self.app.TIME_SLOTS_STR
        ws4.append(jadwal_headers)
        
        # Format header Jadwal
        for col in ws4.iter_cols(min_row=1, max_row=1, max_col=len(jadwal_headers)):
            for cell in col:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.font = openpyxl.styles.Font(bold=True)
        
        # Set column widths
        for ws in [ws1, ws2, ws3, ws4]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Simpan ke buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer


class FileProcessor:
    """Kelas untuk memproses file"""
    
    def __init__(self, app: 'PoliSchedulerApp'):
        self.app = app
        self.error_analyzer = ErrorAnalyzer()
        
    def process_uploaded_file(self, uploaded_file, progress_bar) -> Optional[io.BytesIO]:
        """Proses file yang diupload"""
        try:
            # Validasi file
            progress_bar.progress(10, "Memvalidasi file...")
            validation_result = self._validate_file(uploaded_file)
            
            if not validation_result['is_valid']:
                st.error("❌ " + "\n".join(validation_result['errors']))
                return None
            
            # Analisis error
            progress_bar.progress(20, "Menganalisis struktur...")
            error_reports = self._analyze_sheets(uploaded_file)
            
            # Simpan error reports ke session state
            st.session_state.error_logs = error_reports
            st.session_state.uploaded_file_name = uploaded_file.name
            
            # Tampilkan error jika ada
            has_critical_errors = any(not report['is_valid'] for report in error_reports)
            if has_critical_errors and not self.app.config['auto_fix_errors']:
                st.warning("⚠️ File memiliki error. Aktifkan Auto-fix atau perbaiki manual.")
            
            # Baca data
            progress_bar.progress(40, "Membaca data...")
            df_reguler, df_poleks = self._read_data(uploaded_file)
            
            # Proses jadwal
            progress_bar.progress(60, "Memproses jadwal...")
            data_processor = DataProcessor(self.app)
            
            df_reguler_processed = data_processor.process_schedule(df_reguler, 'Reguler')
            df_poleks_processed = data_processor.process_schedule(df_poleks, 'Poleks')
            
            # Gabungkan hasil
            df_jadwal = pd.concat([df_reguler_processed, df_poleks_processed], 
                                 ignore_index=True)
            
            if df_jadwal.empty:
                st.warning("⚠️ Tidak ada jadwal yang dihasilkan. Periksa data input.")
                return None
            
            # Urutkan
            df_jadwal['HARI_ORDER'] = df_jadwal['HARI'].map(self.app.HARI_ORDER)
            df_jadwal = df_jadwal.sort_values(['POLI ASAL', 'HARI_ORDER', 'DOKTER'])
            df_jadwal = df_jadwal.drop('HARI_ORDER', axis=1).reset_index(drop=True)
            
            # Simpan ke session state
            st.session_state.processed_data = df_jadwal
            st.session_state.visualization_data = df_jadwal
            st.session_state.last_processed_time = datetime.now()
            
            # Buat statistik
            progress_bar.progress(70, "Membuat statistik...")
            time_cols = [col for col in df_jadwal.columns if col in self.app.TIME_SLOTS_STR]
            total_r = (df_jadwal[time_cols] == 'R').sum().sum()
            total_e = (df_jadwal[time_cols] == 'E').sum().sum()
            
            st.session_state.processing_stats = {
                'total_rows': len(df_jadwal),
                'total_reguler': total_r,
                'total_poleks': total_e,
                'poli_count': df_jadwal['POLI ASAL'].nunique(),
                'dokter_count': df_jadwal['DOKTER'].nunique()
            }
            
            # Buat Excel dengan styling
            progress_bar.progress(80, "Membuat file Excel...")
            result_buffer = self._create_styled_excel(uploaded_file, df_jadwal)
            
            progress_bar.progress(95, "Menyelesaikan...")
            return result_buffer
            
        except Exception as e:
            st.error(f"❌ Error dalam memproses file: {str(e)}")
            st.error("Detail error:")
            st.code(traceback.format_exc())
            return None
    
    def _validate_file(self, uploaded_file) -> Dict[str, Any]:
        """Validasi file"""
        try:
            # Cek ekstensi
            if not uploaded_file.name.endswith(('.xlsx', '.xls')):
                return {
                    'is_valid': False,
                    'errors': ["File harus berekstensi .xlsx atau .xls"]
                }
            
            # Cek ukuran file (max 10MB)
            file_size = len(uploaded_file.getvalue()) / (1024 * 1024)
            if file_size > 10:
                return {
                    'is_valid': False,
                    'errors': [f"File terlalu besar ({file_size:.1f}MB). Maksimal 10MB"]
                }
            
            wb = load_workbook(uploaded_file, read_only=True)
            required_sheets = ['Reguler', 'Poleks']
            missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
            
            if missing_sheets:
                return {
                    'is_valid': False,
                    'errors': [f"Sheet berikut tidak ditemukan: {', '.join(missing_sheets)}"]
                }
            
            return {'is_valid': True, 'errors': []}
            
        except Exception as e:
            return {'is_valid': False, 'errors': [f"Error membaca file: {str(e)}"]}
    
    def _analyze_sheets(self, uploaded_file) -> List[Dict[str, Any]]:
        """Analisis sheet untuk error"""
        reports = []
        
        try:
            # Analisis sheet Reguler
            df_reguler = pd.read_excel(uploaded_file, sheet_name='Reguler')
            reguler_report = self.error_analyzer.analyze_excel_structure(df_reguler, 'Reguler')
            reports.append({'sheet': 'Reguler', **reguler_report})
            
            # Analisis sheet Poleks
            df_poleks = pd.read_excel(uploaded_file, sheet_name='Poleks')
            poleks_report = self.error_analyzer.analyze_excel_structure(df_poleks, 'Poleks')
            reports.append({'sheet': 'Poleks', **poleks_report})
            
        except Exception as e:
            reports.append({
                'sheet': 'Unknown',
                'is_valid': False,
                'errors': [f"Error dalam analisis: {str(e)}"]
            })
        
        return reports
    
    def _read_data(self, uploaded_file) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Baca data dari file"""
        df_reguler = pd.read_excel(uploaded_file, sheet_name='Reguler')
        df_poleks = pd.read_excel(uploaded_file, sheet_name='Poleks')
        
        return df_reguler, df_poleks
    
    def _create_styled_excel(self, uploaded_file, df_jadwal: pd.DataFrame) -> io.BytesIO:
        """Buat file Excel dengan styling"""
        new_wb = load_workbook(uploaded_file)
        
        # Hapus sheet Jadwal yang lama jika ada
        if 'Jadwal' in new_wb.sheetnames:
            std = new_wb['Jadwal']
            new_wb.remove(std)
        
        # Buat sheet Jadwal baru
        ws_jadwal = new_wb.create_sheet('Jadwal')
        
        # Tulis header
        headers = ['POLI ASAL', 'JENIS POLI', 'HARI', 'DOKTER'] + self.app.TIME_SLOTS_STR
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_jadwal.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Tulis data
        for row_idx, row_data in enumerate(df_jadwal.to_dict('records'), start=2):
            ws_jadwal.cell(row=row_idx, column=1, value=row_data['POLI ASAL'])
            ws_jadwal.cell(row=row_idx, column=2, value=row_data['JENIS POLI'])
            ws_jadwal.cell(row=row_idx, column=3, value=row_data['HARI'])
            ws_jadwal.cell(row=row_idx, column=4, value=row_data['DOKTER'])
            
            for col_idx, slot in enumerate(self.app.TIME_SLOTS_STR, start=5):
                ws_jadwal.cell(row=row_idx, column=col_idx, value=row_data.get(slot, ''))
        
        # Terapkan styling
        styler = ExcelStyler(self.app)
        styler.apply_styles(ws_jadwal, len(df_jadwal) + 1)
        
        # Simpan ke buffer
        result_buffer = io.BytesIO()
        new_wb.save(result_buffer)
        result_buffer.seek(0)
        
        return result_buffer


# ============================================================================
# TAMPILAN STREAMLIT
# ============================================================================

class StreamlitUI:
    """Kelas untuk mengelola tampilan Streamlit"""
    
    def __init__(self, app: 'PoliSchedulerApp'):
        self.app = app
        self.template_manager = TemplateManager(app)
        self.file_processor = FileProcessor(app)
        self.visualizer = Visualizer(app)
        
    def render_sidebar(self):
        """Render sidebar"""
        with st.sidebar:
            st.title("🏥 Pengisi Jadwal Poli OOP")
            st.markdown("---")
            
            st.subheader("📋 Panduan")
            st.markdown("""
            1. **Download** template untuk format yang benar
            2. **Upload** file Excel yang sudah diisi
            3. **Analisis** error dan perbaiki jika perlu
            4. **Proses** dan **download** hasil
            """)
            
            st.markdown("---")
            
            # Download template
            st.subheader("📥 Template")
            if st.button("Download Template Excel", 
                        type="secondary",
                        use_container_width=True):
                template_buffer = self.template_manager.create_template()
                st.download_button(
                    label="Klik untuk download template",
                    data=template_buffer,
                    file_name="template_jadwal_poli_oop.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            st.markdown("---")
            
            # Auto-fix toggle
            self.app.config['auto_fix_errors'] = st.toggle(
                "🔄 Auto-fix Errors",
                value=self.app.config['auto_fix_errors'],
                help="Secara otomatis memperbaiki error format yang umum"
            )
            
            st.markdown("---")
            
            # Info status
            if st.session_state.processed_data is not None:
                st.subheader("📊 Status Terakhir")
                stats = st.session_state.get('processing_stats', {})
                if stats:
                    st.metric("Total Baris", stats.get('total_rows', 0))
                    st.metric("Jumlah Poli", stats.get('poli_count', 0))
                    st.metric("Jumlah Dokter", stats.get('dokter_count', 0))
            
            st.markdown("---")
            
            # Info kontak/help
            st.caption("❓ Butuh bantuan?")
            st.caption("📧 support@example.com")
    
    def render_main_content(self):
        """Render konten utama"""
        st.title("🏥 Pengisi Jadwal Poli Excel - OOP Version")
        st.caption("Aplikasi berbasis OOP untuk mengisi jadwal poli secara otomatis")
        
        # Tabs
        tab1, tab2, tab3, tab4 = st.tabs([
            "📤 Upload & Proses", 
            "🔍 Error Analyzer", 
            "📊 Visualisasi", 
            "⚙️ Pengaturan"
        ])
        
        with tab1:
            self.render_upload_tab()
        
        with tab2:
            self.render_error_analyzer_tab()
        
        with tab3:
            self.render_visualization_tab()
        
        with tab4:
            self.render_settings_tab()
    
    def render_upload_tab(self):
        """Render tab upload"""
        col1, col2 = st.columns([2, 1])
        
        with col1:
            uploaded_file = st.file_uploader(
                "Upload file Excel (.xlsx)", 
                type=['xlsx'],
                help="Upload file dengan format yang sesuai template",
                key="file_uploader"
            )
        
        with col2:
            if uploaded_file:
                file_size = len(uploaded_file.getvalue()) / 1024
                st.metric("📏 Ukuran File", f"{file_size:.1f} KB")
                
                try:
                    excel_file = pd.ExcelFile(uploaded_file)
                    sheet_count = len(excel_file.sheet_names)
                    st.success(f"✅ {sheet_count} sheet ditemukan")
                    
                    # Tampilkan sheet names
                    with st.expander("Lihat sheet"):
                        for sheet in excel_file.sheet_names:
                            st.caption(f"• {sheet}")
                except Exception as e:
                    st.error(f"❌ File tidak valid: {e}")
        
        if uploaded_file:
            st.markdown("---")
            
            # Preview sheet
            with st.expander("📄 Preview Sheet", expanded=False):
                sheet_names = pd.ExcelFile(uploaded_file).sheet_names
                selected_sheet = st.selectbox("Pilih sheet untuk preview:", sheet_names)
                
                if selected_sheet:
                    try:
                        df_preview = pd.read_excel(uploaded_file, sheet_name=selected_sheet, nrows=10)
                        st.dataframe(df_preview, width='stretch')
                        st.caption(f"Menampilkan 10 dari {len(pd.read_excel(uploaded_file, sheet_name=selected_sheet))} baris")
                    except Exception as e:
                        st.warning(f"Tidak dapat membaca sheet {selected_sheet}: {e}")
            
            # Tombol proses
            st.markdown("---")
            col_proses1, col_proses2, col_proses3 = st.columns([1, 2, 1])
            with col_proses2:
                if st.button("🚀 Proses Jadwal", 
                           type="primary", 
                           use_container_width=True):
                    
                    with st.spinner("Memproses..."):
                        progress_bar = st.progress(0)
                        
                        result_buffer = self.file_processor.process_uploaded_file(
                            uploaded_file, progress_bar
                        )
                        
                        if result_buffer:
                            progress_bar.progress(100)
                            st.success("✅ File berhasil diproses!")
                            
                            # Tampilkan statistik
                            if st.session_state.processing_stats:
                                stats = st.session_state.processing_stats
                                col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                                with col_stat1:
                                    st.metric("Total Baris", stats['total_rows'])
                                with col_stat2:
                                    st.metric("Slot Reguler", stats['total_reguler'])
                                with col_stat3:
                                    st.metric("Slot Poleks", stats['total_poleks'])
                                with col_stat4:
                                    st.metric("Jumlah Poli", stats['poli_count'])
                            
                            # Tombol download
                            st.download_button(
                                label="📥 Download Hasil Excel",
                                data=result_buffer,
                                file_name=f"jadwal_hasil_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                type="primary"
                            )
    
    def render_error_analyzer_tab(self):
        """Render tab error analyzer"""
        st.subheader("🔍 Error Analyzer")
        
        if 'error_logs' in st.session_state and st.session_state.error_logs:
            for error_report in st.session_state.error_logs:
                with st.expander(f"📋 Laporan untuk sheet: {error_report['sheet']}", expanded=True):
                    report_html = ErrorAnalyzer.create_error_report(error_report)
                    st.markdown(report_html, unsafe_allow_html=True)
                    
                    # Tampilkan rekomendasi perbaikan
                    if error_report['errors'] or error_report['warnings']:
                        st.markdown("### 🛠️ Rekomendasi Perbaikan:")
                        
                        if error_report['errors']:
                            st.markdown("#### ❌ Error yang perlu diperbaiki:")
                            for error in error_report['errors'][:5]:
                                if "kolom yang diperlukan" in error.lower():
                                    st.markdown("- **Tambahkan kolom yang hilang** sesuai template")
                                elif "format waktu" in error.lower():
                                    st.markdown("- **Perbaiki format waktu** menggunakan format `HH.MM - HH.MM` atau `HH:MM-HH:MM`")
                                elif "file tidak valid" in error.lower():
                                    st.markdown("- **Gunakan template** yang disediakan")
                        
                        if error_report['warnings']:
                            st.markdown("#### ⚠️ Peringatan:")
                            for warning in error_report['warnings']:
                                if "kosong" in warning.lower():
                                    st.markdown("- **Hapus baris kosong** atau isi data")
                                elif "duplikat" in warning.lower():
                                    st.markdown("- **Periksa duplikat data** dokter dan jadwal")
                        
                        if self.app.config['auto_fix_errors']:
                            st.success("✅ Auto-fix akan memperbaiki error format waktu secara otomatis")
        else:
            st.info("ℹ️ Belum ada data error yang dianalisis. Upload file terlebih dahulu.")
            
            # Contoh error
            with st.expander("📝 Contoh format yang benar"):
                st.markdown("""
                ### Format waktu yang diterima:
                - `08.00 - 10.30`
                - `08:00 - 10:30`
                - `08.00-10.30`
                - `08:00-10:30`
                
                ### Format yang TIDAK diterima:
                - `8-10` (tanpa menit)
                - `08.00 s/d 10.30` (menggunakan 's/d')
                - `08:00 sampai 10:30` (menggunakan 'sampai')
                - `08.00-10` (format tidak lengkap)
                """)
    
    def render_visualization_tab(self):
        """Render tab visualisasi"""
        st.subheader("📊 Visualisasi Jadwal")
        
        if (st.session_state.processed_data is not None and 
            not st.session_state.processed_data.empty):
            
            df = st.session_state.processed_data
            
            # Pilihan visualisasi
            viz_type = st.selectbox(
                "Pilih jenis visualisasi:",
                ["Heatmap", "Gantt Chart", "Statistik", "Tabel Interaktif", "Ringkasan"]
            )
            
            if viz_type == "Heatmap":
                fig = self.visualizer.create_heatmap(df)
                if fig:
                    st.plotly_chart(fig, width='stretch')
                    
                    # Legenda
                    with st.expander("📖 Legenda"):
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.markdown("⬜ **Kosong** - Tidak ada jadwal")
                        with col2:
                            st.markdown("🟩 **Reguler** - Jadwal reguler")
                        with col3:
                            st.markdown("🟦 **Poleks** - Jadwal poleks")
                        st.markdown("🟥 **Over Limit** - Melebihi batas maksimal")
            
            elif viz_type == "Gantt Chart":
                fig = self.visualizer.create_schedule_gantt(df)
                if fig:
                    st.plotly_chart(fig, width='stretch')
                else:
                    st.warning("Data tidak cukup untuk Gantt Chart")
            
            elif viz_type == "Statistik":
                fig_pie, fig_bar = self.visualizer.create_statistics_charts(df)
                
                if fig_pie:
                    col1, col2 = st.columns(2)
                    with col1:
                        st.plotly_chart(fig_pie, width='stretch')
                    with col2:
                        if fig_bar:
                            st.plotly_chart(fig_bar, width='stretch')
                        else:
                            st.info("Tidak ada data untuk bar chart")
                    
                    # Tampilkan statistik detail
                    st.markdown("### 📈 Statistik Detail")
                    time_cols = [col for col in df.columns if col in self.app.TIME_SLOTS_STR]
                    total_r = (df[time_cols] == 'R').sum().sum()
                    total_e = (df[time_cols] == 'E').sum().sum()
                    total_slots = len(df) * len(time_cols)
                    
                    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                    with col_stat1:
                        st.metric("Total Baris", len(df))
                    with col_stat2:
                        st.metric("Total Slot", total_slots)
                    with col_stat3:
                        percent_r = (total_r/total_slots*100) if total_slots > 0 else 0
                        st.metric("Slot Reguler", total_r, f"{percent_r:.1f}%")
                    with col_stat4:
                        percent_e = (total_e/total_slots*100) if total_slots > 0 else 0
                        st.metric("Slot Poleks", total_e, f"{percent_e:.1f}%")
                else:
                    st.warning("Tidak ada data untuk statistik")
            
            elif viz_type == "Tabel Interaktif":
                # Filter data
                col_filter1, col_filter2, col_filter3 = st.columns(3)
                with col_filter1:
                    selected_poli = st.multiselect(
                        "Filter Poli:",
                        options=df['POLI ASAL'].unique(),
                        default=df['POLI ASAL'].unique()[:3] if len(df['POLI ASAL'].unique()) > 3 else df['POLI ASAL'].unique()
                    )
                with col_filter2:
                    selected_hari = st.multiselect(
                        "Filter Hari:",
                        options=df['HARI'].unique(),
                        default=df['HARI'].unique()
                    )
                with col_filter3:
                    selected_jenis = st.multiselect(
                        "Filter Jenis:",
                        options=df['JENIS POLI'].unique(),
                        default=df['JENIS POLI'].unique()
                    )
                
                # Terapkan filter
                if not selected_poli:
                    selected_poli = df['POLI ASAL'].unique()
                if not selected_hari:
                    selected_hari = df['HARI'].unique()
                if not selected_jenis:
                    selected_jenis = df['JENIS POLI'].unique()
                
                filtered_df = df[
                    df['POLI ASAL'].isin(selected_poli) &
                    df['HARI'].isin(selected_hari) &
                    df['JENIS POLI'].isin(selected_jenis)
                ].copy()
                
                # Format tabel dengan warna
                def color_cells(val):
                    if val == 'R':
                        return 'background-color: green; color: white'
                    elif val == 'E':
                        return 'background-color: blue; color: white'
                    elif val == '':
                        return 'background-color: #f0f0f0'
                    else:
                        return ''
                
                # Pilih kolom untuk ditampilkan
                time_cols = [col for col in filtered_df.columns if col in self.app.TIME_SLOTS_STR]
                display_cols = ['POLI ASAL', 'JENIS POLI', 'HARI', 'DOKTER'] + time_cols
                
                # Tampilkan tabel
                st.dataframe(
                    filtered_df[display_cols].style.applymap(color_cells, subset=time_cols),
                    width='stretch',
                    height=400
                )
                
                # Download filtered data
                if not filtered_df.empty:
                    csv = filtered_df.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="📥 Download Data Tersaring (CSV)",
                        data=csv,
                        file_name="jadwal_tersaring.csv",
                        mime="text/csv",
                        width='stretch'
                    )
            
            elif viz_type == "Ringkasan":
                # Ringkasan statistik
                st.markdown("### 📋 Ringkasan Jadwal")
                
                time_cols = [col for col in df.columns if col in self.app.TIME_SLOTS_STR]
                
                # Statistik per poli
                st.markdown("#### 📊 Per Poli")
                poli_stats = df.groupby('POLI ASAL').agg({
                    'DOKTER': 'nunique',
                    'HARI': lambda x: x.nunique()
                }).rename(columns={'DOKTER': 'Jml Dokter', 'HARI': 'Jml Hari'})
                
                # Tambahkan jumlah slot
                poli_stats['Slot Reguler'] = df.groupby('POLI ASAL').apply(
                    lambda x: (x[time_cols] == 'R').sum().sum()
                )
                poli_stats['Slot Poleks'] = df.groupby('POLI ASAL').apply(
                    lambda x: (x[time_cols] == 'E').sum().sum()
                )
                
                st.dataframe(poli_stats, width='stretch')
                
                # Statistik per hari
                st.markdown("#### 📅 Per Hari")
                hari_stats = []
                for hari in self.app.HARI_INDONESIA:
                    hari_df = df[df['HARI'] == hari]
                    if not hari_df.empty:
                        hari_stats.append({
                            'Hari': hari,
                            'Jml Poli': hari_df['POLI ASAL'].nunique(),
                            'Jml Dokter': hari_df['DOKTER'].nunique(),
                            'Slot Reguler': (hari_df[time_cols] == 'R').sum().sum(),
                            'Slot Poleks': (hari_df[time_cols] == 'E').sum().sum()
                        })
                
                if hari_stats:
                    st.dataframe(pd.DataFrame(hari_stats), width='stretch')
                
                # Waktu tersibuk
                st.markdown("#### ⏰ Waktu Tersibuk")
                slot_usage = {}
                for slot in time_cols:
                    slot_r = (df[slot] == 'R').sum()
                    slot_e = (df[slot] == 'E').sum()
                    slot_usage[slot] = {'Reguler': slot_r, 'Poleks': slot_e, 'Total': slot_r + slot_e}
                
                slot_df = pd.DataFrame(slot_usage).T.sort_values('Total', ascending=False)
                st.dataframe(slot_df.head(10), width='stretch')
        
        else:
            st.info("ℹ️ Belum ada data yang diproses. Upload dan proses file terlebih dahulu.")
            
            # Preview contoh visualisasi
            with st.expander("👁️ Preview Visualisasi"):
                st.markdown("""
                ### Fitur Visualisasi:
                
                1. **Heatmap** - Peta panas distribusi jadwal
                2. **Gantt Chart** - Timeline jadwal per dokter
                3. **Statistik** - Grafik distribusi slot
                4. **Tabel Interaktif** - Filter dan ekspor data
                5. **Ringkasan** - Statistik detail per poli dan hari
                """)
    
    def render_settings_tab(self):
        """Render tab pengaturan"""
        st.subheader("⚙️ Pengaturan Aplikasi")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### ⏰ Pengaturan Waktu")
            
            # Jam mulai
            col_hour, col_min = st.columns(2)
            with col_hour:
                new_start_hour = st.number_input(
                    "Jam mulai",
                    min_value=5,
                    max_value=12,
                    value=self.app.config['start_hour'],
                    help="Jam mulai jadwal (5-12)"
                )
            with col_min:
                new_start_minute = st.number_input(
                    "Menit mulai",
                    min_value=0,
                    max_value=59,
                    value=self.app.config['start_minute'],
                    step=5,
                    help="Menit mulai jadwal"
                )
            
            # Interval
            new_interval = st.selectbox(
                "Interval waktu (menit)",
                options=[15, 20, 30, 45, 60],
                index=[15, 20, 30, 45, 60].index(self.app.config['interval_minutes'])
                if self.app.config['interval_minutes'] in [15, 20, 30, 45, 60] else 2
            )
            
            # Batasan
            new_max_poleks = st.number_input(
                "Batas maksimal Poleks per slot",
                min_value=1,
                max_value=20,
                value=self.app.config['max_poleks_per_slot'],
                help="Jika lebih dari batas ini, akan ditandai merah"
            )
        
        with col2:
            st.markdown("### 🎨 Pengaturan Tampilan")
            
            # Tema warna
            new_color_theme = st.selectbox(
                "Tema warna visualisasi",
                ["Default", "Hijau-Biru", "Merah-Kuning", "Pastel", "High Contrast"],
                index=["Default", "Hijau-Biru", "Merah-Kuning", "Pastel", "High Contrast"]
                .index(self.app.config['color_theme'])
                if self.app.config['color_theme'] in ["Default", "Hijau-Biru", "Merah-Kuning", "Pastel", "High Contrast"]
                else 0
            )
            
            # Auto-save
            auto_save = st.toggle(
                "Simpan pengaturan secara otomatis",
                value=True,
                help="Simpan pengaturan untuk sesi berikutnya"
            )
            
            # Preview time slots
            st.markdown("### 👁️ Preview Time Slots")
            
            # Regenerate time slots untuk preview
            preview_slots = []
            current_time = time(new_start_hour, new_start_minute)
            end_time = time(14, 30)
            
            while current_time <= end_time:
                preview_slots.append(current_time.strftime("%H:%M"))
                current_datetime = datetime.combine(datetime.today(), current_time)
                next_datetime = current_datetime + timedelta(minutes=new_interval)
                current_time = next_datetime.time()
            
            # Tampilkan preview
            cols = st.columns(min(4, len(preview_slots)))
            for i, slot in enumerate(preview_slots):
                with cols[i % len(cols)]:
                    st.info(f"**{i+1}.** {slot}")
        
        # Tombol apply
        if st.button("💾 Terapkan Pengaturan", type="primary", width='stretch'):
            # Cek perubahan
            config_changed = (
                new_start_hour != self.app.config['start_hour'] or
                new_start_minute != self.app.config['start_minute'] or
                new_interval != self.app.config['interval_minutes'] or
                new_max_poleks != self.app.config['max_poleks_per_slot'] or
                new_color_theme != self.app.config['color_theme']
            )
            
            if config_changed:
                # Update config
                self.app.config.update({
                    'start_hour': new_start_hour,
                    'start_minute': new_start_minute,
                    'interval_minutes': new_interval,
                    'max_poleks_per_slot': new_max_poleks,
                    'color_theme': new_color_theme
                })
                
                # Update time slots
                self.app._update_time_slots()
                
                st.success("✅ Pengaturan berhasil diterapkan!")
                st.session_state.config_changed = True
                
                # Reset processed data jika perlu
                if st.session_state.processed_data is not None:
                    st.warning("⚠️ Pengaturan berubah. Proses ulang file untuk menerapkan perubahan.")
            else:
                st.info("ℹ️ Tidak ada perubahan pengaturan.")
        
        # Reset button
        if st.button("🔄 Reset ke Default", type="secondary", width='stretch'):
            self.app.initialize_constants()
            st.success("✅ Pengaturan telah direset ke default")
            st.rerun()


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Fungsi utama"""
    # Inisialisasi aplikasi
    app = PoliSchedulerApp()
    
    # Inisialisasi UI
    ui = StreamlitUI(app)
    
    # Render aplikasi
    ui.render_sidebar()
    ui.render_main_content()


if __name__ == "__main__":
    main()
