import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import time, timedelta, datetime, date
import io
import re
import json
from typing import Dict, List, Optional

# Konfigurasi halaman
st.set_page_config(
    page_title="Pengisi Jadwal Poli Excel + Kanban",
    page_icon="🏥📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Warna untuk sel
FILL_R = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
FILL_E = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
FILL_OVERLIMIT = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Warna untuk Kanban
KANBAN_COLORS = {
    "todo": "#FF6B6B",
    "in_progress": "#FFD93D",
    "review": "#6BCF7F",
    "done": "#4D96FF"
}

# Konfigurasi default
TIME_SLOTS = [
    time(7, 30), time(8, 0), time(8, 30), time(9, 0), time(9, 30),
    time(10, 0), time(10, 30), time(11, 0), time(11, 30), time(12, 0),
    time(12, 30), time(13, 0), time(13, 30), time(14, 0), time(14, 30)
]
TIME_SLOTS_STR = [t.strftime("%H:%M") for t in TIME_SLOTS]

HARI_ORDER = {"Senin": 1, "Selasa": 2, "Rabu": 3, "Kamis": 4, "Jum'at": 5}
HARI_INDONESIA = ["Senin", "Selasa", "Rabu", "Kamis", "Jum'at"]

# ==================== SESSION STATE INITIALIZATION ====================

# Inisialisasi session state untuk Kanban
def init_session_state():
    """Initialize all session state variables"""
    if 'kanban_tasks' not in st.session_state:
        st.session_state.kanban_tasks = {
            "todo": [],
            "in_progress": [],
            "review": [],
            "done": []
        }
    
    if 'kanban_next_id' not in st.session_state:
        st.session_state.kanban_next_id = 1
    
    if 'processed_result' not in st.session_state:
        st.session_state.processed_result = None
    
    if 'processed_data' not in st.session_state:
        st.session_state.processed_data = None
    
    if 'processed_filename' not in st.session_state:
        st.session_state.processed_filename = ""
    
    if 'editing_task' not in st.session_state:
        st.session_state.editing_task = None

# Panggil fungsi init
init_session_state()

# ==================== UTILITY FUNCTIONS ====================

def parse_time_range(time_str):
    """Parse rentang waktu dari string seperti '08.00 - 10.30'"""
    if pd.isna(time_str) or str(time_str).strip() == "":
        return None, None
    
    clean_str = str(time_str).strip().replace(' ', '').replace('.', ':')
    pattern = r'(\d{1,2}:\d{2})-(\d{1,2}:\d{2})'
    match = re.search(pattern, clean_str)
    if not match:
        return None, None
    
    try:
        start_str, end_str = match.groups()
        start_hour, start_minute = map(int, start_str.split(':'))
        end_hour, end_minute = map(int, end_str.split(':'))
        
        start_time = time(start_hour, start_minute)
        end_time = time(end_hour, end_minute)
        
        if end_time < start_time:
            end_time = time(end_hour + 24, end_minute)
        
        return start_time, end_time
    except:
        return None, None

def time_overlap(slot_start, slot_end, schedule_start, schedule_end):
    """Cek apakah slot waktu overlap dengan jadwal"""
    if schedule_start is None or schedule_end is None:
        return False
    
    return not (slot_end <= schedule_start or slot_start >= schedule_end)

def process_schedule(df, jenis_poli):
    """Proses dataframe jadwal"""
    results = []
    
    for (dokter, poli_asal), group in df.groupby(['Nama Dokter', 'Poli Asal']):
        hari_schedules = {}
        
        for hari in HARI_INDONESIA:
            if hari not in group.columns:
                continue
                
            time_ranges = []
            for time_str in group[hari]:
                start_time, end_time = parse_time_range(time_str)
                if start_time and end_time:
                    if end_time > time(14, 30):
                        end_time = time(14, 30)
                    time_ranges.append((start_time, end_time))
            
            hari_schedules[hari] = time_ranges
        
        for hari in HARI_INDONESIA:
            if hari not in hari_schedules or not hari_schedules[hari]:
                continue
                
            merged_ranges = []
            for start, end in sorted(hari_schedules[hari]):
                if not merged_ranges:
                    merged_ranges.append([start, end])
                else:
                    last_start, last_end = merged_ranges[-1]
                    if start <= last_end:
                        merged_ranges[-1][1] = max(last_end, end)
                    else:
                        merged_ranges.append([start, end])
            
            row = {
                'POLI ASAL': poli_asal,
                'JENIS POLI': jenis_poli,
                'HARI': hari,
                'DOKTER': dokter
            }
            
            for i, slot_time in enumerate(TIME_SLOTS):
                slot_start = slot_time
                slot_end = (datetime.combine(datetime.today(), slot_start) + 
                           timedelta(minutes=30)).time()
                
                has_overlap = any(
                    time_overlap(slot_start, slot_end, start, end)
                    for start, end in merged_ranges
                )
                
                if has_overlap:
                    row[TIME_SLOTS_STR[i]] = 'R' if jenis_poli == 'Reguler' else 'E'
                else:
                    row[TIME_SLOTS_STR[i]] = ''
            
            results.append(row)
    
    return pd.DataFrame(results)

def apply_styles(ws, max_row):
    """Terapkan styling ke sheet Jadwal"""
    e_counts = {hari: {slot: 0 for slot in TIME_SLOTS_STR} for hari in HARI_INDONESIA}
    
    for row in range(2, max_row + 1):
        hari = ws.cell(row=row, column=3).value
        if hari not in HARI_INDONESIA:
            continue
            
        for col_idx, slot in enumerate(TIME_SLOTS_STR, start=5):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value == 'E':
                e_counts[hari][slot] += 1
    
    for row in range(2, max_row + 1):
        hari = ws.cell(row=row, column=3).value
        if hari not in HARI_INDONESIA:
            continue
            
        for col_idx, slot in enumerate(TIME_SLOTS_STR, start=5):
            cell = ws.cell(row=row, column=col_idx)
            
            if cell.value == 'R':
                cell.fill = FILL_R
            elif cell.value == 'E':
                cell.fill = FILL_E
                
                if e_counts[hari][slot] > 7:
                    e_rows_for_slot = []
                    for r in range(2, max_row + 1):
                        if (ws.cell(row=r, column=3).value == hari and 
                            ws.cell(row=r, column=col_idx).value == 'E'):
                            e_rows_for_slot.append(r)
                    
                    if len(e_rows_for_slot) > 7:
                        if row in e_rows_for_slot[7:]:
                            cell.fill = FILL_OVERLIMIT

# ==================== KANBAN FUNCTIONS ====================

def render_kanban_board():
    """Render Kanban board"""
    st.subheader("📋 Kanban Board - Tracking Jadwal")
    
    # Statistics row
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    with col_stat1:
        total_tasks = sum(len(tasks) for tasks in st.session_state.kanban_tasks.values())
        st.metric("Total Tasks", total_tasks)
    with col_stat2:
        st.metric("Todo", len(st.session_state.kanban_tasks["todo"]))
    with col_stat3:
        st.metric("In Progress", len(st.session_state.kanban_tasks["in_progress"]))
    with col_stat4:
        st.metric("Done", len(st.session_state.kanban_tasks["done"]))
    
    # Filter controls
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        filter_priority = st.selectbox(
            "Filter Priority",
            ["All", "high", "medium", "low"],
            key="filter_priority"
        )
    with col2:
        # Get all assignees
        all_assignees = set()
        for column in st.session_state.kanban_tasks.values():
            for task in column:
                if task.get("assignee"):
                    all_assignees.add(task["assignee"])
        
        filter_assignee = st.selectbox(
            "Filter Assignee",
            ["All"] + sorted(list(all_assignees)),
            key="filter_assignee"
        )
    with col3:
        # Get all tags
        all_tags = set()
        for column in st.session_state.kanban_tasks.values():
            for task in column:
                for tag in task.get("tags", []):
                    all_tags.add(tag)
        
        filter_tags = st.multiselect(
            "Filter Tags",
            options=sorted(list(all_tags)),
            key="filter_tags"
        )
    
    st.markdown("---")
    
    # Add new task section
    with st.expander("➕ Tambah Task Baru", expanded=False):
        add_task_form()
    
    st.markdown("---")
    
    # Create Kanban columns
    columns = st.columns(4)
    column_names = ["Todo", "In Progress", "Review", "Done"]
    column_keys = ["todo", "in_progress", "review", "done"]
    
    for idx, (col, col_name, col_key) in enumerate(zip(columns, column_names, column_keys)):
        with col:
            # Column header
            st.markdown(
                f"""
                <div style='
                    background-color: {KANBAN_COLORS[col_key]};
                    padding: 10px;
                    border-radius: 5px;
                    margin-bottom: 10px;
                    color: white;
                    font-weight: bold;
                    text-align: center;
                '>
                {col_name} ({len(st.session_state.kanban_tasks[col_key])})
                </div>
                """,
                unsafe_allow_html=True
            )
            
            # Tasks in this column
            tasks = st.session_state.kanban_tasks[col_key]
            
            # Apply filters
            filtered_tasks = tasks
            if filter_priority != "All":
                filtered_tasks = [t for t in filtered_tasks if t.get("priority", "medium") == filter_priority]
            if filter_assignee != "All":
                filtered_tasks = [t for t in filtered_tasks if t.get("assignee") == filter_assignee]
            if filter_tags:
                filtered_tasks = [t for t in filtered_tasks if any(tag in t.get("tags", []) for tag in filter_tags)]
            
            if not filtered_tasks:
                st.info("No tasks in this column")
            else:
                for task in filtered_tasks:
                    render_task_card(task, col_key)

def render_task_card(task, current_column):
    """Render individual task card"""
    priority_colors = {
        "high": "#FF0000",
        "medium": "#FFA500",
        "low": "#008000"
    }
    
    priority = task.get("priority", "medium")
    border_color = priority_colors.get(priority, "#FFA500")
    
    with st.container():
        st.markdown(
            f"""
            <div style='
                background-color: white;
                border: 2px solid {border_color};
                border-radius: 8px;
                padding: 12px;
                margin-bottom: 10px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            '>
            <div style='display: flex; justify-content: space-between; align-items: center;'>
                <strong>#{task.get("id", "?")}</strong>
                <small style='color: {border_color}; font-weight: bold;'>{priority.upper()}</small>
            </div>
            <h4 style='margin: 8px 0; color: #333;'>{task.get("title", "No Title")}</h4>
            """,
            unsafe_allow_html=True
        )
        
        if task.get("description"):
            st.caption(task.get("description"))
        
        # Tags
        if task.get("tags") and len(task.get("tags", [])) > 0:
            tags_html = ""
            for tag in task.get("tags", []):
                tags_html += f"<span style='background-color: #e0e0e0; padding: 2px 6px; border-radius: 10px; font-size: 0.8em; margin-right: 4px;'>{tag}</span>"
            st.markdown(tags_html, unsafe_allow_html=True)
        
        # Assignee and due date
        col1, col2 = st.columns(2)
        with col1:
            if task.get("assignee"):
                st.caption(f"👤 {task.get('assignee')}")
        with col2:
            if task.get("due_date"):
                st.caption(f"📅 {task.get('due_date')}")
        
        st.markdown("</div>", unsafe_allow_html=True)
        
        # Task actions
        column_keys = ["todo", "in_progress", "review", "done"]
        current_idx = column_keys.index(current_column)
        
        col_move1, col_move2, col_delete = st.columns(3)
        
        # Move left button
        with col_move1:
            if current_idx > 0:
                if st.button("⬅️", key=f"left_{task['id']}_{current_column}", help="Move left"):
                    move_task(task['id'], current_column, column_keys[current_idx - 1])
                    st.rerun()
        
        # Move right button
        with col_move2:
            if current_idx < 3:
                if st.button("➡️", key=f"right_{task['id']}_{current_column}", help="Move right"):
                    move_task(task['id'], current_column, column_keys[current_idx + 1])
                    st.rerun()
        
        # Delete button
        with col_delete:
            if st.button("🗑️", key=f"delete_{task['id']}_{current_column}", help="Delete"):
                delete_task(task['id'], current_column)
                st.rerun()

def add_task_form():
    """Form untuk menambah task baru"""
    with st.form(key="add_task_form"):
        title = st.text_input("Judul Task*", placeholder="Contoh: Review jadwal Poli Anak", key="task_title")
        description = st.text_area("Deskripsi", placeholder="Detail task...", key="task_desc")
        
        col1, col2 = st.columns(2)
        with col1:
            priority = st.selectbox("Priority", ["low", "medium", "high"], key="task_priority")
            assignee = st.text_input("Assignee", placeholder="Nama penanggung jawab", key="task_assignee")
        with col2:
            due_date = st.date_input("Due Date", min_value=date.today(), key="task_due_date")
            tags_input = st.text_input("Tags (pisahkan dengan koma)", placeholder="jadwal, review, poli", key="task_tags")
        
        submitted = st.form_submit_button("➕ Tambah Task", use_container_width=True)
        
        if submitted:
            if not title.strip():
                st.error("Judul task harus diisi!")
                return
                
            # Parse tags
            tags = []
            if tags_input:
                tags = [tag.strip() for tag in tags_input.split(",") if tag.strip()]
            
            new_task = {
                "id": st.session_state.kanban_next_id,
                "title": title,
                "description": description,
                "priority": priority,
                "due_date": due_date.strftime("%Y-%m-%d") if due_date else None,
                "assignee": assignee,
                "tags": tags,
                "created_by": "User",
                "created_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
            }
            
            st.session_state.kanban_tasks["todo"].append(new_task)
            st.session_state.kanban_next_id += 1
            st.success(f"Task '{title}' ditambahkan!")
            st.rerun()

def move_task(task_id, from_column, to_column):
    """Pindah task antar column"""
    task_to_move = None
    for task in st.session_state.kanban_tasks[from_column]:
        if task["id"] == task_id:
            task_to_move = task
            break
    
    if task_to_move:
        st.session_state.kanban_tasks[from_column].remove(task_to_move)
        task_to_move["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        st.session_state.kanban_tasks[to_column].append(task_to_move)

def delete_task(task_id, from_column):
    """Hapus task"""
    st.session_state.kanban_tasks[from_column] = [
        task for task in st.session_state.kanban_tasks[from_column] 
        if task["id"] != task_id
    ]

def export_kanban_to_excel():
    """Export Kanban board ke Excel"""
    wb = openpyxl.Workbook()
    
    # Buat sheet untuk setiap column
    for column_name in ["todo", "in_progress", "review", "done"]:
        ws = wb.create_sheet(title=column_name.capitalize())
        ws.append(["ID", "Title", "Description", "Priority", "Due Date", 
                  "Assignee", "Tags", "Created By", "Created Date", "Last Updated"])
        
        for task in st.session_state.kanban_tasks[column_name]:
            ws.append([
                task.get("id"),
                task.get("title"),
                task.get("description"),
                task.get("priority"),
                task.get("due_date"),
                task.get("assignee"),
                ", ".join(task.get("tags", [])),
                task.get("created_by"),
                task.get("created_date"),
                task.get("last_updated")
            ])
    
    # Hapus sheet default
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def import_kanban_from_excel(file):
    """Import Kanban dari Excel"""
    try:
        wb = load_workbook(file)
        imported_data = {}
        
        for sheet_name in wb.sheetnames:
            column_key = sheet_name.lower()
            if column_key in ["todo", "in_progress", "review", "done"]:
                df = pd.read_excel(file, sheet_name=sheet_name)
                # Convert dataframe to list of dicts
                tasks = []
                for _, row in df.iterrows():
                    task = {
                        "id": int(row.get("ID", 0)),
                        "title": str(row.get("Title", "")),
                        "description": str(row.get("Description", "")),
                        "priority": str(row.get("Priority", "medium")),
                        "due_date": str(row.get("Due Date", "")) if pd.notna(row.get("Due Date")) else None,
                        "assignee": str(row.get("Assignee", "")) if pd.notna(row.get("Assignee")) else "",
                        "tags": [tag.strip() for tag in str(row.get("Tags", "")).split(",") if tag.strip()],
                        "created_by": str(row.get("Created By", "System")),
                        "created_date": str(row.get("Created Date", "")),
                        "last_updated": str(row.get("Last Updated", ""))
                    }
                    tasks.append(task)
                
                imported_data[column_key] = tasks
        
        # Update session state
        for column, tasks in imported_data.items():
            st.session_state.kanban_tasks[column] = tasks
        
        # Update next_id
        max_id = 0
        for column in st.session_state.kanban_tasks.values():
            for task in column:
                if isinstance(task, dict) and "id" in task:
                    max_id = max(max_id, task["id"])
        st.session_state.kanban_next_id = max_id + 1
        
        return True
    except Exception as e:
        st.error(f"Error importing: {str(e)}")
        return False

def create_sample_kanban_tasks():
    """Buat sample tasks untuk testing"""
    sample_tasks = {
        "todo": [
            {
                "id": 1,
                "title": "Review Jadwal Poli Anak",
                "description": "Periksa jadwal reguler dan poleks untuk Poli Anak",
                "priority": "high",
                "due_date": (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d"),
                "assignee": "Koordinator Poli",
                "tags": ["jadwal", "review", "anak"],
                "created_by": "System",
                "created_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
            },
            {
                "id": 2,
                "title": "Update Template Excel",
                "description": "Update template untuk format baru",
                "priority": "medium",
                "due_date": (datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d"),
                "assignee": "Admin",
                "tags": ["template", "excel"],
                "created_by": "System",
                "created_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
            }
        ],
        "in_progress": [
            {
                "id": 3,
                "title": "Proses Jadwal Bulan Desember",
                "description": "Memproses jadwal dokter untuk bulan Desember",
                "priority": "high",
                "due_date": (datetime.now() + timedelta(days=2)).strftime("%Y-%m-%d"),
                "assignee": "Staff IT",
                "tags": ["processing", "desember"],
                "created_by": "System",
                "created_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
            }
        ],
        "review": [
            {
                "id": 4,
                "title": "Validasi Jadwal Poli Bedah",
                "description": "Validasi jadwal dokter bedah",
                "priority": "medium",
                "due_date": (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d"),
                "assignee": "Manager",
                "tags": ["validation", "bedah"],
                "created_by": "System",
                "created_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
            }
        ],
        "done": [
            {
                "id": 5,
                "title": "Setup Aplikasi Jadwal",
                "description": "Setup aplikasi jadwal poli",
                "priority": "low",
                "due_date": datetime.now().strftime("%Y-%m-%d"),
                "assignee": "Developer",
                "tags": ["setup", "app"],
                "created_by": "System",
                "created_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
            }
        ]
    }
    
    st.session_state.kanban_tasks = sample_tasks
    st.session_state.kanban_next_id = 6

# ==================== PROCESS FILE FUNCTION ====================

def process_file(uploaded_file):
    """Proses file Excel dan kembalikan buffer"""
    try:
        # Baca file
        wb = load_workbook(uploaded_file)
        
        # Cek sheet yang diperlukan
        required_sheets = ['Reguler', 'Poleks', 'Jadwal']
        missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
        
        if missing_sheets:
            raise ValueError(f"Sheet berikut tidak ditemukan: {', '.join(missing_sheets)}")
        
        # Baca data
        df_reguler = pd.read_excel(uploaded_file, sheet_name='Reguler')
        df_poleks = pd.read_excel(uploaded_file, sheet_name='Poleks')
        
        # Proses jadwal
        df_reguler_processed = process_schedule(df_reguler, 'Reguler')
        df_poleks_processed = process_schedule(df_poleks, 'Poleks')
        
        # Gabungkan hasil
        df_jadwal = pd.concat([df_reguler_processed, df_poleks_processed], ignore_index=True)
        
        # Urutkan
        df_jadwal['HARI_ORDER'] = df_jadwal['HARI'].map(HARI_ORDER)
        df_jadwal = df_jadwal.sort_values(['POLI ASAL', 'HARI_ORDER', 'DOKTER', 'JENIS POLI'])
        df_jadwal = df_jadwal.drop('HARI_ORDER', axis=1)
        
        # Reset index
        df_jadwal = df_jadwal.reset_index(drop=True)
        
        # Simpan ke session state untuk preview
        st.session_state.processed_data = df_jadwal
        
        # Buat workbook baru
        new_wb = load_workbook(uploaded_file)
        
        # Hapus sheet Jadwal yang lama jika ada
        if 'Jadwal' in new_wb.sheetnames:
            std = new_wb['Jadwal']
            new_wb.remove(std)
        
        # Buat sheet Jadwal baru
        ws_jadwal = new_wb.create_sheet('Jadwal')
        
        # Tulis header
        headers = ['POLI ASAL', 'JENIS POLI', 'HARI', 'DOKTER'] + TIME_SLOTS_STR
        for col_idx, header in enumerate(headers, start=1):
            ws_jadwal.cell(row=1, column=col_idx, value=header)
        
        # Tulis data
        for row_idx, row_data in enumerate(df_jadwal.to_dict('records'), start=2):
            ws_jadwal.cell(row=row_idx, column=1, value=row_data['POLI ASAL'])
            ws_jadwal.cell(row=row_idx, column=2, value=row_data['JENIS POLI'])
            ws_jadwal.cell(row=row_idx, column=3, value=row_data['HARI'])
            ws_jadwal.cell(row=row_idx, column=4, value=row_data['DOKTER'])
            
            for col_idx, slot in enumerate(TIME_SLOTS_STR, start=5):
                ws_jadwal.cell(row=row_idx, column=col_idx, value=row_data.get(slot, ''))
        
        # Terapkan styling
        apply_styles(ws_jadwal, len(df_jadwal) + 1)
        
        # Simpan ke buffer
        result_buffer = io.BytesIO()
        new_wb.save(result_buffer)
        result_buffer.seek(0)
        
        # Simpan ke session state
        st.session_state.processed_result = result_buffer
        st.session_state.processed_filename = uploaded_file.name
        
        return result_buffer
        
    except Exception as e:
        raise Exception(f"Error dalam memproses file: {str(e)}")

# ==================== MAIN APP ====================

def main():
    # Sidebar
    with st.sidebar:
        st.title("🏥📋 Jadwal Poli + Kanban")
        st.markdown("---")
        
        # Navigation
        page = st.radio(
            "Navigasi",
            ["📤 Upload & Proses", "📋 Kanban Board", "📊 Analytics", "⚙️ Settings"],
            index=0
        )
        
        st.markdown("---")
        
        if page == "📤 Upload & Proses":
            st.subheader("📋 Panduan")
            st.markdown("""
            1. **Upload** file Excel
            2. **Proses** jadwal
            3. **Download** hasil
            4. **Track** di Kanban
            """)
            
        elif page == "📋 Kanban Board":
            st.subheader("Kanban Actions")
            
            # Add sample data button
            if st.button("🔄 Add Sample Data", use_container_width=True):
                create_sample_kanban_tasks()
                st.success("Sample data added!")
                st.rerun()
            
            # Export/Import Kanban
            st.markdown("### Import/Export")
            col_exp, col_imp = st.columns(2)
            with col_exp:
                if st.button("📤 Export", use_container_width=True):
                    buffer = export_kanban_to_excel()
                    st.download_button(
                        label="Download Excel",
                        data=buffer,
                        file_name="kanban_board.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            with col_imp:
                kanban_file = st.file_uploader(
                    "Import from Excel",
                    type=['xlsx'],
                    key="kanban_import",
                    label_visibility="collapsed"
                )
                if kanban_file:
                    if import_kanban_from_excel(kanban_file):
                        st.success("Kanban imported successfully!")
                        st.rerun()
            
            # Quick stats
            st.markdown("---")
            total_tasks = sum(len(tasks) for tasks in st.session_state.kanban_tasks.values())
            st.metric("Total Tasks", total_tasks)
            
        st.markdown("---")
        st.caption("v1.0 • Dengan fitur Kanban")
    
    # Main content based on selected page
    if page == "📤 Upload & Proses":
        render_upload_page()
    elif page == "📋 Kanban Board":
        render_kanban_board()
    elif page == "📊 Analytics":
        render_analytics_page()
    elif page == "⚙️ Settings":
        render_settings_page()

def render_upload_page():
    """Halaman upload dan proses"""
    st.title("🏥 Pengisi Jadwal Poli Excel")
    st.caption("Dengan fitur Kanban untuk tracking progress")
    
    # Upload file section
    uploaded_file = st.file_uploader(
        "Upload file Excel (.xlsx)", 
        type=['xlsx'],
        help="Upload file dengan format yang sesuai",
        key="file_uploader"
    )
    
    if uploaded_file:
        col1, col2 = st.columns([2, 1])
        with col2:
            file_size = len(uploaded_file.getvalue()) / 1024
            st.metric("File Size", f"{file_size:.1f} KB")
        
        # Preview
        with st.expander("📄 Preview File", expanded=False):
            sheet_names = pd.ExcelFile(uploaded_file).sheet_names
            st.write(f"**Sheets:** {', '.join(sheet_names)}")
            
            selected_sheet = st.selectbox("Pilih sheet untuk preview:", sheet_names, key="sheet_preview")
            if selected_sheet:
                try:
                    df_preview = pd.read_excel(uploaded_file, sheet_name=selected_sheet, nrows=5)
                    st.dataframe(df_preview, use_container_width=True)
                except:
                    st.warning(f"Tidak dapat membaca sheet {selected_sheet}")
        
        # Proses button - TOMBOL UTAMA
        if st.button("🚀 Proses Jadwal & Buat Task Kanban", type="primary", use_container_width=True):
            with st.spinner("Memproses file..."):
                try:
                    # Proses file
                    result_buffer = process_file(uploaded_file)
                    
                    # Auto-create Kanban task untuk tracking
                    auto_create_kanban_task(uploaded_file.name)
                    
                    st.success("✅ File berhasil diproses dan task Kanban dibuat!")
                    
                except Exception as e:
                    st.error(f"❌ Error: {str(e)}")
        
        # Tampilkan tombol download jika sudah ada hasil yang diproses
        if st.session_state.processed_result is not None:
            st.markdown("---")
            st.subheader("📥 Download Hasil")
            
            col_dl1, col_dl2, col_dl3 = st.columns([1, 2, 1])
            with col_dl2:
                # Tombol download untuk hasil jadwal
                st.download_button(
                    label=f"📥 Download Hasil Jadwal",
                    data=st.session_state.processed_result,
                    file_name=f"jadwal_hasil_{st.session_state.processed_filename}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_jadwal"
                )
            
            # Tampilkan preview data yang diproses
            with st.expander("👁️ Preview Data yang Diproses", expanded=False):
                if st.session_state.processed_data is not None:
                    st.dataframe(st.session_state.processed_data.head(20), use_container_width=True)
                    
                    # Statistik
                    col_stat1, col_stat2, col_stat3 = st.columns(3)
                    with col_stat1:
                        total_rows = len(st.session_state.processed_data)
                        st.metric("Total Baris", total_rows)
                    with col_stat2:
                        total_r = (st.session_state.processed_data[TIME_SLOTS_STR] == 'R').sum().sum()
                        st.metric("Slot Reguler", total_r)
                    with col_stat3:
                        total_e = (st.session_state.processed_data[TIME_SLOTS_STR] == 'E').sum().sum()
                        st.metric("Slot Poleks", total_e)
            
            # Tombol untuk reset/clear hasil
            if st.button("🗑️ Clear Hasil", type="secondary", use_container_width=True):
                st.session_state.processed_result = None
                st.session_state.processed_data = None
                st.session_state.processed_filename = ""
                st.rerun()
    
    else:
        # Tampilkan info jika belum upload file
        st.info("👆 Silakan upload file Excel untuk memulai")
        
        # Template download
        with st.expander("📥 Download Template File", expanded=False):
            st.markdown("Download template untuk format yang benar:")
            
            # Buat template sederhana
            if st.button("Buat Template Excel", key="create_template"):
                wb = openpyxl.Workbook()
                
                # Sheet Poli Asal
                ws1 = wb.active
                ws1.title = "Poli Asal"
                ws1.append(["No", "Nama Poli", "kode sheet"])
                
                # Sheet Reguler
                ws2 = wb.create_sheet("Reguler")
                ws2.append(["Nama Dokter", "Poli Asal", "Jenis Poli", "Senin", "Selasa", "Rabu", "Kamis", "Jum'at"])
                ws2.append(["dr. Contoh Dokter, Sp.A", "Poli Anak", "Reguler", "08.00 - 10.30", "", "", "", ""])
                
                # Sheet Poleks
                ws3 = wb.create_sheet("Poleks")
                ws3.append(["Nama Dokter", "Poli Asal", "Jenis Poli", "Senin", "Selasa", "Rabu", "Kamis", "Jum'at"])
                ws3.append(["dr. Contoh Dokter, Sp.A", "Poli Anak", "Poleks", "07.30 - 08.25", "", "", "", ""])
                
                # Sheet Jadwal
                ws4 = wb.create_sheet("Jadwal")
                ws4.append(["POLI ASAL", "JENIS POLI", "HARI", "DOKTER"] + TIME_SLOTS_STR)
                
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                st.download_button(
                    label="Download Template.xlsx",
                    data=buffer,
                    file_name="template_jadwal_poli.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

def auto_create_kanban_task(filename):
    """Otomatis buat task Kanban setelah proses"""
    new_task = {
        "id": st.session_state.kanban_next_id,
        "title": f"Review Jadwal: {filename}",
        "description": f"File jadwal yang diproses: {filename}. Perlu direview oleh koordinator.",
        "priority": "medium",
        "due_date": (datetime.now() + timedelta(days=2)).strftime("%Y-%m-%d"),
        "assignee": "Koordinator Poli",
        "tags": ["jadwal", "review", "auto-generated"],
        "created_by": "System",
        "created_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    
    st.session_state.kanban_tasks["todo"].append(new_task)
    st.session_state.kanban_next_id += 1

def render_analytics_page():
    """Halaman analytics"""
    st.title("📊 Analytics & Reports")
    
    # Kanban analytics
    col1, col2, col3, col4 = st.columns(4)
    
    total_tasks = sum(len(tasks) for tasks in st.session_state.kanban_tasks.values())
    done_tasks = len(st.session_state.kanban_tasks["done"])
    
    with col1:
        st.metric("Total Tasks", total_tasks)
    with col2:
        st.metric("Tasks Done", done_tasks)
    with col3:
        progress = (done_tasks / total_tasks * 100) if total_tasks > 0 else 0
        st.metric("Completion", f"{progress:.1f}%")
    with col4:
        st.metric("Active Tasks", len(st.session_state.kanban_tasks["in_progress"]))
    
    # Task distribution
    st.subheader("Task Distribution")
    
    # Simple bar chart dengan HTML/CSS
    column_names = ["Todo", "In Progress", "Review", "Done"]
    column_keys = ["todo", "in_progress", "review", "done"]
    task_counts = [len(st.session_state.kanban_tasks[key]) for key in column_keys]
    
    # Buat chart sederhana
    chart_html = """
    <div style="margin: 20px 0; padding: 20px; background-color: #f8f9fa; border-radius: 10px;">
    """
    
    max_count = max(task_counts) if task_counts else 1
    
    for name, count, color in zip(column_names, task_counts, KANBAN_COLORS.values()):
        width = (count / max_count * 100) if max_count > 0 else 0
        chart_html += f"""
        <div style="margin-bottom: 15px;">
            <div style="display: flex; justify-content: space-between; margin-bottom: 5px;">
                <span><strong>{name}</strong></span>
                <span>{count} tasks</span>
            </div>
            <div style="width: 100%; background-color: #e0e0e0; border-radius: 5px; height: 25px;">
                <div style="width: {width}%; background-color: {color}; height: 25px; border-radius: 5px; 
                         transition: width 0.5s;"></div>
            </div>
        </div>
        """
    
    chart_html += "</div>"
    st.markdown(chart_html, unsafe_allow_html=True)
    
    # Recent activity
    st.subheader("Recent Activity")
    
    # Kumpulkan semua task dan urutkan berdasarkan last_updated
    all_tasks = []
    for column, tasks in st.session_state.kanban_tasks.items():
        for task in tasks:
            task['column'] = column
            all_tasks.append(task)
    
    # Sort by last_updated (newest first)
    all_tasks_sorted = sorted(
        all_tasks, 
        key=lambda x: x.get('last_updated', ''), 
        reverse=True
    )[:10]  # Ambil 10 terbaru
    
    if all_tasks_sorted:
        for task in all_tasks_sorted:
            with st.container():
                st.markdown(f"""
                <div style="padding: 10px; margin-bottom: 10px; border-left: 4px solid {KANBAN_COLORS[task['column']]}; 
                         background-color: white; border-radius: 5px;">
                    <strong>#{task['id']} - {task['title']}</strong><br>
                    <small>Status: {task['column'].replace('_', ' ').title()} • 
                    Updated: {task.get('last_updated', 'N/A')}</small>
                </div>
                """, unsafe_allow_html=True)
    else:
        st.info("No recent activity")

def render_settings_page():
    """Halaman settings"""
    st.title("⚙️ Settings")
    
    # Kanban settings
    with st.expander("Kanban Settings", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            auto_create = st.checkbox("Auto-create task setelah proses file", value=True)
            notify_done = st.checkbox("Notify when task moves to Done", value=True)
        with col2:
            enable_deadlines = st.checkbox("Enable task deadlines", value=True)
            show_tags = st.checkbox("Show tags on task cards", value=True)
    
    # Schedule settings
    with st.expander("Schedule Settings"):
        col1, col2, col3 = st.columns(3)
        with col1:
            start_time = st.time_input("Start time", value=time(7, 30))
        with col2:
            end_time = st.time_input("End time", value=time(14, 30))
        with col3:
            interval = st.selectbox("Time interval (minutes)", [15, 30, 60], index=1)
    
    # Data management
    with st.expander("Data Management", expanded=False):
        st.warning("⚠️ Hati-hati dengan operasi ini!")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🔄 Reset Kanban Board", use_container_width=True, key="reset_kanban"):
                st.session_state.kanban_tasks = {
                    "todo": [], "in_progress": [], "review": [], "done": []
                }
                st.session_state.kanban_next_id = 1
                st.success("Kanban board telah direset!")
                st.rerun()
        
        with col2:
            if st.button("🗑️ Clear All Data", use_container_width=True, key="clear_all"):
                st.session_state.kanban_tasks = {
                    "todo": [], "in_progress": [], "review": [], "done": []
                }
                st.session_state.kanban_next_id = 1
                st.session_state.processed_result = None
                st.session_state.processed_data = None
                st.session_state.processed_filename = ""
                st.success("Semua data telah dihapus!")
                st.rerun()
    
    # Export settings
    if st.button("💾 Export All Settings", use_container_width=True, key="export_settings"):
        settings = {
            "kanban_settings": {
                "auto_create": True,
                "notify_done": True,
                "enable_deadlines": True,
                "show_tags": True
            },
            "schedule_settings": {
                "start_time": start_time.strftime("%H:%M"),
                "end_time": end_time.strftime("%H:%M"),
                "interval": interval
            }
        }
        
        st.download_button(
            label="Download Settings.json",
            data=json.dumps(settings, indent=2),
            file_name="kanban_settings.json",
            mime="application/json"
        )

if __name__ == "__main__":
    main()
